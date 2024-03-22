import sys
from PyQt5 import QtCore,QtGui,QtWidgets
from PyQt5.uic import loadUi
from PyQt5.QtCore import QDate , QBuffer, QByteArray , QTime
from PyQt5.QtGui import QImage,QPixmap 
from PyQt5.QtGui import QIcon
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QIODevice , QUrl, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QWidget ,QCompleter,QApplication ,QMainWindow,QStackedWidget,QGraphicsDropShadowEffect, QCalendarWidget , QBoxLayout
from PyQt5.QtWidgets import QMessageBox,QLabel,QTableWidgetItem
from PyQt5.QtWidgets import QDialog
import hashlib
import sqlite3
import os
import datetime
from bs4 import BeautifulSoup
import requests
from urllib3.exceptions import InsecureRequestWarning
from PyQt5.QtCore import Qt
import re
from users import Ui_Dialog
import json
import uuid  # Importar el módulo uuid para generar identificadores únicos
import openpyxl
from openpyxl.drawing.image import Image
class IngresoUsuario(QMainWindow):
    def __init__(self):
        super(IngresoUsuario, self).__init__()
        loadUi("./interfaces/loggin.ui", self)
        self.setWindowTitle("Login")
        self.btn_login.clicked.connect(self.ingreso)
        self.bt_salir.clicked.connect(self.salida)
        self.ingresoAnterior()
        
    
    def ingresoAnterior(self):
        usuario, contrasena = self.cargar_datos_acceso()
        if usuario and contrasena:
            self.label_2.setText(f"Hola nuevamente {usuario}")
            self.txt_username.setText(usuario)
            self.txt_password.setText(contrasena)
        else:
            print("No se encontraron datos de acceso almacenados.")
    def salida(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()
            
    
    def guardar_datos_acceso(self,usuario, contrasena):
        datos_acceso = {"usuario": usuario, "contrasena": contrasena}
        with open("datos_acceso.json", "w") as archivo:
            json.dump(datos_acceso, archivo)

    # Función para cargar los datos de acceso desde un archivo
    def cargar_datos_acceso(self):
        try:
            with open("datos_acceso.json", "r") as archivo:
                datos_acceso = json.load(archivo)
                return datos_acceso["usuario"], datos_acceso["contrasena"]
        except FileNotFoundError:
            return None, None
    def cifrar_contrasenia(self, contrasenia):
        # Cifrar la contraseña usando un algoritmo de hash (SHA-256 en este caso)
        cifrado = hashlib.sha256()
        cifrado.update(contrasenia.encode('utf-8'))
        return cifrado.hexdigest()

    def get_greeting_message(self, nombre):
        hora_actual = datetime.datetime.now().time()
        if datetime.time(5, 0, 0) <= hora_actual < datetime.time(12, 0, 0):
            return f"Buenos días {nombre}\n¿Qué deseas hacer hoy?"
        elif datetime.time(12, 0, 0) <= hora_actual < datetime.time(18, 0, 0):
            return f"Buenas tardes {nombre}\n¿Qué deseas hacer hoy?"
        elif datetime.time(18, 0, 0) <= hora_actual or hora_actual < datetime.time(5, 0, 0):
            return f"Buenas noches {nombre}\n¿Qué deseas hacer hoy?"
        else:
            return f"Hola {nombre}\n¿Qué deseas hacer hoy?"

    def authenticate_user(self, username, password):
        if not username or not password:
            QMessageBox.warning(self, "Error", "Por favor ingrese usuario y contraseña.")
            return

        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        cursor.execute("SELECT * FROM Users WHERE Username = ?", (username,))
        usuario = cursor.fetchone()
        conexion.close()

        if usuario:
            contrasenia_cifrada_ingresada = self.cifrar_contrasenia(password)
            contrasenia_cifrada_almacenada = usuario[1]  # Suponiendo que el hash se almacena en el segundo campo de la tabla
            if contrasenia_cifrada_ingresada == contrasenia_cifrada_almacenada:
                text_for_menu = self.get_greeting_message(username)
                id_user = usuario[2]
                self.open_menu_principal(text_for_menu, id_user)
                
                
            else:
                QMessageBox.warning(self, "Error", "Contraseña Incorrecta.")
        else:
            QMessageBox.warning(self, "Error", "Nombre de usuario no encontrado.")

    def open_menu_principal(self, text_for_menu, id_user):
        menu_principal = MenuPrincipal(id_user)
        menu_principal.lb_nombre.setText(text_for_menu)
        
        # Establecer la ventana en modo de pantalla completa
        menu_principal.showMaximized()
       
        menu_principal.setWindowTitle("Menu Principal")

     
        # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
        widget.addWidget(menu_principal)
       
        widget.setCurrentIndex(widget.currentIndex() + 1)

        self.close()

    def ingreso(self):
        nombre = self.txt_username.text()
        password = self.txt_password.text()
        self.authenticate_user(nombre, password)
        self.guardar_datos_acceso(nombre, password)
        print("Datos de acceso guardados exitosamente.")
   
class Registro(QMainWindow):
    def __init__(self, id_user):
        super(Registro, self).__init__()
        loadUi("interfaces/dogtores.ui", self)
        self.id_user = id_user  # Nuevo atributo para almacenar el id_user
        self.btn_agg.clicked.connect(self.registrarUsuario)
        self.btn_clear.clicked.connect(self.clearInputs)
        self.actionSalir.triggered.connect(self.close)
        self.actionvolver_edit.triggered.connect(self.backLogin)
        self.bt_photo.clicked.connect(self.addPhoto)
        self.in_cedula.textChanged.connect(self.verificar_existencia_cedula)
        self.in_mail.editingFinished.connect(self.mostrar_mensaje_mail)
        self.in_number.textChanged.connect(self.mostrar_mensaje_telefono)
        self.users_dialog = None  # Definir users_dialog como un atributo de la clase
        
    def ingresoRegistro(self):
        if self.id_user is not None:
            # Verificar el tipo de usuario antes de permitir el acceso al registro
            if self.validar_tipo_usuario(self.id_user, ["Administrador", "Doctor"]):
                registroview = Registro(self.id_user)
                widget.addWidget(registroview)
                widget.setCurrentIndex(widget.currentIndex() + 1)
                registroview.show()
                self.hide()
            else:
                QMessageBox.warning(self, "Acceso denegado", "Solo Administradores y Doctores pueden acceder al registro.")
        else:
            QMessageBox.warning(self, "Error", "ID de usuario no válido. Inicie sesión primero.")

    def validar_tipo_usuario(self, id_user, tipos_permitidos):
        # Función para validar si el tipo de usuario es permitido
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE Id_User = ?", (id_user,))
        tipo_usuario = cursor.fetchone()
        conexion.close()

        return tipo_usuario[0] in tipos_permitidos if tipo_usuario else False
    def backLogin(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea volver al menú de edición de usuario?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            edicion = EditDoctor(self.id_user)
            widget.addWidget(edicion)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            edicion.show()
            self.hide()
            
    def addPhoto(self):
        filenames, _ = QFileDialog.getOpenFileNames(self, "Seleccionar imágenes", "", "Archivos de imagen (*.png *.jpg *.bmp)")

        if len(filenames) >= 1:
            pixmap1 = QPixmap(filenames[0])
            self.foto.setPixmap(pixmap1)
        else:
            QMessageBox.information(self, "Imagenes", "Por favor, selecciona una imagen")

    def validar_correo_electronico(self, correo):
        patron = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        return re.match(patron, correo) is not None

    def mostrar_mensaje_mail(self):
        correo = self.in_mail.text()
        if self.validar_correo_electronico(correo):
            self.btn_agg.setEnabled(True)
        else:
            QMessageBox.warning(self, "Correo invalido", "Por favor introduzca un correo valido")
            self.btn_agg.setEnabled(False)
            return

    def validar_numeros(self, cadena):
        patron = r'^[0-9]+$'
        return re.match(patron, cadena) is not None

    def mostrar_mensaje_telefono(self):
        numero = self.in_number.text()
        if self.validar_numeros(numero):
            self.btn_agg.setEnabled(True)
        else:
            QMessageBox.information(self, "Solo numeros", "Número inválido")
            self.btn_agg.setEnabled(False)
            return

    def clearInputs(self):
        self.in_cedula.clear()
        self.in_name.clear()
        self.in_apell.clear()
        self.in_age.clear()
        self.in_mail.clear()
        self.in_number.clear()
        self.in_dir.clear()
        self.btn_m.setChecked(False)
        self.btn_f.setChecked(False)
        self.in_espec.clear()
        self.foto.clear()

    def cifrar_contrasenia(self, contrasenia):
        cifrado = hashlib.sha256()
        cifrado.update(contrasenia.encode('utf-8'))
        return cifrado.hexdigest()

    def verificar_existencia_cedula(self):
        cedula = self.in_cedula.text()

        if not cedula:
            return

        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        cursor.execute('SELECT * FROM Users WHERE Cedula = ?', (cedula,))
        existe_cedula = cursor.fetchone() is not None
        conexion.close()

        if existe_cedula:
            QMessageBox.warning(self, "Error", "Ya existe alguien con la misma cédula.")
            self.btn_agg.setEnabled(False)
            return
        else:
            self.btn_agg.setEnabled(True)

    def close(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QtWidgets.QApplication.quit()

    def usuario_existe(self, username):
        # Conectar a la base de datos
        conexion = sqlite3.connect('interfaces/database.db')
        # Crear un cursor
        cursor = conexion.cursor()
        # Consultar si el usuario ya existe
        cursor.execute('SELECT COUNT(*) FROM Users WHERE Username = ?', (username,))
        resultado = cursor.fetchone()[0]
        # Cerrar la conexión
        conexion.close()
        # Si resultado es mayor que 0, significa que el usuario ya existe
        return resultado > 0
    
    def validar_cedula(self, cedula):
        cedula_pattern = re.compile(r'^\d{8}$')  # Asume que la cédula debe contener 8 a 10 dígitos
        if not cedula_pattern.match(cedula):
            QMessageBox.warning(self, "Error", "Ingrese una cédula válida (8 numeros).")
            return True
        return False

    def validar_edad(self, edad):
        edad_pattern = re.compile(r'^[1-9]\d*$')  # Asume que la edad debe ser un número entero positivo
        if not edad_pattern.match(edad):
            QMessageBox.warning(self, "Error", "Ingrese una edad válida (solo números).")
            return True
        return False
    def registrarUsuario(self):
        cedula = self.in_cedula.text()
        nombre = self.in_name.text()
        apellido = self.in_apell.text()
        edad = self.in_age.text()
        mail = self.in_mail.text()
        valor_sexo = ""
        foto = self.foto.pixmap()
        if self.btn_m.isChecked():
            valor_sexo = "Masculino"
        elif self.btn_f.isChecked():
            valor_sexo = "Femenino"

        if not cedula or not nombre or not apellido or not edad or not valor_sexo or not mail:
            QMessageBox.critical(self, "Error", "Por favor, complete todos los campos básicos.")
            return

        if self.validar_cedula(cedula) or self.validar_edad(edad):
            return

        if len(cedula) < 8:
            QMessageBox.critical(self, "Error", "La cédula debe tener mínimo 8 caracteres.")
            return

        self.datos_basicos = {
            'cedula': cedula,
            'nombre': nombre,
            'apellido': apellido,
            'edad': edad,
            'sexo': valor_sexo,
            'mail': mail,
        }

        # Mostrar diálogo para completar información adicional
        self.openUsersDialog()

    def guardarDatosUsuarios(self):
        username = self.users_dialog.in_user.text()
        password = self.users_dialog.in_password.text()
        passwordRepeat = self.users_dialog.in_conf.text()
        telefono = self.in_number.text()
        direccion = self.in_dir.text()
        especialidad = self.in_espec.text()
        foto = self.foto.pixmap()
        tipoUser = None
        if self.users_dialog.tipo_admin.isChecked():
            tipoUser = "Administrador"
        elif self.users_dialog.tipo_doc.isChecked():
            tipoUser = "Doctor"
        elif self.users_dialog.tipo_user.isChecked():
            tipoUser = "Usuario"
        else:
            tipoUser = "Tipo no seleccionado"

        if not username or not password or not passwordRepeat or not telefono or not direccion or not especialidad:
            QMessageBox.critical(self, "Error", "Por favor, complete todos los campos de usuario.")
            return

        if len(username) < 6:
            QMessageBox.critical(self, "Error", "Su nombre de usuario debe tener mínimo 6 caracteres")
            return

        if len(password) < 8:
            QMessageBox.critical(self, "Error", "La contraseña debe tener mínimo 8 caracteres.")
            return

        if password != passwordRepeat:
            QMessageBox.warning(self, "Error", "Las contraseñas no coinciden.")
            return

        if not foto:
            QMessageBox.warning(self, "Advertencia", "Debes importar una imagen antes de guardar.")
            return

        cedula = self.datos_basicos['cedula']
        nombre = self.datos_basicos['nombre']
        apellido = self.datos_basicos['apellido']
        edad = self.datos_basicos['edad']
        valor_sexo = self.datos_basicos['sexo']
        mail = self.datos_basicos['mail']

        # Validar si el nombre de usuario ya existe
        if self.usuario_existe(username):
            QMessageBox.warning(self, "Error", "El nombre de usuario ya existe. \nIngrese uno distinto")
            return

        # Cifrar la contraseña
        contrasenia_cifrada = self.cifrar_contrasenia(password)

        foto_image = foto.toImage()
        foto_bytes = QByteArray()
        buffer = QBuffer(foto_bytes)
        buffer.open(QIODevice.WriteOnly)
        foto_image.save(buffer, "PNG")
        foto_bytes = buffer.data()
        buffer.close()

        # Conectar a la base de datos
        conexion = sqlite3.connect('interfaces/database.db')
        # Crear un cursor
        cursor = conexion.cursor()
        # Insertar datos en la tabla Users
        cursor.execute(
            'INSERT INTO Users (Username, Password, Cedula, Nombres, Apellidos, Sexo, Edad, Direccion, Telefono, Mail, Especialidad, Imagen, Tipo) VALUES (?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (username, contrasenia_cifrada, cedula, nombre, apellido, valor_sexo, edad, direccion, telefono, mail,
            especialidad, foto_bytes, tipoUser))
        # Confirmar la transacción
        conexion.commit()
        # Cerrar la conexión
        conexion.close()

        QMessageBox.information(self, "Éxito", "Registro exitoso")

        # Cerrar el diálogo de usuarios
        self.users_dialog.close()
        
    def limpiar(self):
        self.users_dialog.in_user.clear()
        self.users_dialog.in_password.clear()
        self.users_dialog.in_conf.clear()

        
    def openUsersDialog(self):
        # Crear una instancia del diálogo y asignarla como un atributo de la instancia de la clase
        self.users_dialog = QDialog(self)
        ui = Ui_Dialog()
        ui.setupUi(self.users_dialog)

        # Conectar la señal de aceptar en el diálogo de usuarios a una función
        ui.btn_agg.clicked.connect(self.guardarDatosUsuarios)
        ui.btn_clear.clicked.connect(self.limpiar)

        # Acceder directamente a los atributos necesarios en el diálogo de usuarios
        self.users_dialog.in_user = self.users_dialog.findChild(QtWidgets.QLineEdit, 'in_user')
        self.users_dialog.in_password = self.users_dialog.findChild(QtWidgets.QLineEdit, 'in_password')
        self.users_dialog.in_conf = self.users_dialog.findChild(QtWidgets.QLineEdit, 'in_conf')
        self.users_dialog.tipo_admin = self.users_dialog.findChild(QtWidgets.QRadioButton, 'bt_admin')
        self.users_dialog.tipo_doc = self.users_dialog.findChild(QtWidgets.QRadioButton, 'bt_doc')
        self.users_dialog.tipo_user = self.users_dialog.findChild(QtWidgets.QRadioButton, 'bt_user')
        
        self.users_dialog.exec_()
class Ui_montos(QMainWindow):
    def __init__(self, id_user):
        super(Ui_montos, self).__init__()
        loadUi("./interfaces/montos.ui", self)
        self.setWindowTitle("Montos")
        self.id_user=id_user
        self.tratamientos = {
            "Triaje": ["Consulta e Historia Clínica sin informe", "Consulta e Historia Clínica con informe"],
            "Periodoncia": ["Tartectomía y pulido simple (1 sesión)", "Tartectomía y pulido simple (2-3 sesiones)","Aplicación tópica de fluór","Cirguia periodontal (por cuadrante)"],
            "Blanqueamiento": ["Blanqueamiento intrapulpar", "Blanquemaineto maxilar superior e inferior (2 sesiones de 20 min c/u)"],
            "Operatoria": ["Obturaciones provisionales","Obturaciones con Amalgama","Obturaciones con vidrio ionomerico pequeña","Obturaciones con vidrio ionomerico grande","Obturaciones con resina fotocurada"],
            "Endodoncia": ["Pulpotomías formocreasoladas","Emergencias Endodontica","Tratamiento endodontico monoradicular","Tratamiento endodontico biradicular","Tratamiento endodontico multiradicular","Desobturación conductos"],
            "Radiografias Periaciales": ["Adultos e infantes"],
            "Cirugias": ["Exodoncia simple","Exodoncia quirurgica","Exodoncia de dientes temporales","Exodoncia de corales erupcionadas/incluidas"],
            "Protesis": ["Coronas provisionales por unidad","Muñon artificial monoradicular","Muñon artificial multiradicular","Incrustacion resina/metálica","Unidad de corona meta-porcelana","Cementado de protesis fija"],
            "Protesis removibles metalicas y/o acrilicas": ["1 a 3 unidades","4 a 6 unidades","7 a 12 unidades","Unidadad adicional","Ganchos contorneados retentativas acrilicas c/u","Reparaciones protesis acrilicas y/oo agregar un diente a la protesis"],
            "Protesis totales": ["Dentadura superior o inferior (incluye controles post-inatalción) c/u"],
            "Implantes dentales": ["Honorarios cirujano por implante","Implante y aditamientos","Injertos óseos (1cc)","PRF (incluye bionalista y extraccion de sangre + centrifugado)","Corona metal porcelana sobre implante","DPR acrilica"],
        }

        self.btn_agg.clicked.connect(self.aggMontos)
        self.btn_clear.clicked.connect(self.clear)
        self.actionSalir.triggered.connect(self.salir)
        self.actionRegresar_al_menu_prinicipal.triggered.connect(self.backmenu)

        self.t_0 = self.findChild(QtWidgets.QComboBox, "t_0")
        self.t_0.addItem("Seleccione el tipo de honorario")
        self.t_0.addItems(list(self.tratamientos.keys()))

        self.t_0.currentTextChanged.connect(self.loadTratamientos)

        self.loadTratamientos()
        self.btn_saveC.clicked.connect(self.guardarMonto)
        
    def guardarMonto(self):
        texto = self.invalor_dia.text()
        if texto and re.match(r'^-?\d*\.?\d+$', texto):
            valor_dola = float(texto)  
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            
            cursor.execute("SELECT ID_V FROM Valor")
            registro_existente = cursor.fetchone()
            
            if registro_existente:
                cursor.execute("UPDATE Valor SET valorDola = ? WHERE ID_V = ?", (valor_dola, registro_existente[0]))
            else:
                cursor.execute('INSERT INTO Valor (ID_V, valorDola) VALUES (NULL, ?)', (valor_dola,))
            
            conexion.commit()
            conexion.close()
            QMessageBox.warning(None, "Aviso", "El monto ingresado fue guardado con éxito")
            self.invalor_dia.clear()
        elif not texto:
            QMessageBox.warning(None, "Error", "Por favor, rellene el campo.")
        else:
            QMessageBox.warning(None, "Error", "Por favor, ingrese solo números en el campo de monto.")
            self.invalor_dia.clear()    


        
    def backmenu(self):        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)

            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")

            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
            
    def loadTratamientos(self):
        self.clear()
        seleccion_t_0 = self.t_0.currentText()
        if seleccion_t_0 == "Seleccione el tipo de honorario":
            return

        tratamientos_honorario = self.tratamientos.get(seleccion_t_0, [])

        for i, tratamiento in enumerate(tratamientos_honorario, start=1):
            getattr(self, f't_{i}').setText(tratamiento)

    def aggMontos(self):
        tipo_tratamiento = self.t_0.currentText()

        tratamientos_montos = []
        num_montos_introducidos = 0
        
        for i in range(1, 7):
            tratamiento = getattr(self, f't_{i}').text()
            monto = getattr(self, f'monto_{i}').text()

            # Verifica si el tratamiento y el monto no están vacíos
            if tratamiento and monto:
                if not re.match(r'^-?\d*\.?\d+$', monto):
                    QMessageBox.warning(self, "Error", "Por favor, ingrese solo números en el campo de monto.")
                    self.clear()
                    return  

                
                tratamientos_montos.append((tratamiento, monto))
                num_montos_introducidos += 1  # Incrementa el contador
        if num_montos_introducidos == 0:
            QtWidgets.QMessageBox.warning(None, 'Error', 'No se han introducido montos')
            return
        
        
        montos=QMessageBox.question(self, "Cantidad de Montos",
                                    f"Se han introducido {num_montos_introducidos} montos.\n¿Deseas seguir con la ejecución?",
                                    QMessageBox.Yes | QMessageBox.No
                                     )

        if montos == QMessageBox.Yes:
            if len(tipo_tratamiento) <= 0:
                QMessageBox.warning(self, "Advertencia", "Debes ingresar el tipo de tratamiento")
                return

            try:
                # Conecta a la base de datos y almacena los datos
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()

                # Verifica si el tratamiento ya existe en la base de datos
                for tratamiento, monto in tratamientos_montos:
                    cursor.execute("SELECT monto FROM Trata WHERE tipo_tratamiento = ? AND tratamiento = ?", (tipo_tratamiento, tratamiento))
                    existing_record = cursor.fetchone()

                    if existing_record:
                        # Si existe el tratamiento, preguntar al usuario si desea actualizar el monto
                        update = QMessageBox.question(self, "Tratamiento Existente",
                                                    f"El tratamiento '{tratamiento}' ya existe con un monto de '{existing_record[0]}'. ¿Desea actualizar el monto?",
                                                    QMessageBox.Yes | QMessageBox.No)

                        if update == QMessageBox.Yes:
                            cursor.execute("UPDATE Trata SET monto = ? WHERE tipo_tratamiento = ? AND tratamiento = ?",
                                        (monto, tipo_tratamiento, tratamiento))
                            QMessageBox.information(self, "Éxito", f"¡Monto actualizado para el tratamiento '{tratamiento}'!")
                    else:
                        # Si no existe, insertar el nuevo tratamiento y monto
                        cursor.execute("INSERT INTO Trata (tipo_tratamiento, tratamiento, monto) VALUES (?, ?, ?)",
                                    (tipo_tratamiento, tratamiento, monto))
                        QMessageBox.information(self, "Éxito", "Datos de tratamientos añadidos correctamente")

                for i in range(1, 7):
                    getattr(self, f't_{i}').clear()
                    getattr(self, f"monto_{i}").clear()

                conexion.commit()
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Error", "Error al insertar o actualizar datos en la base de datos: " + str(e))
            finally:
                conexion.close()
        else:
            self.clear()

    def clear(self):
        for i in range(1, 7):
            getattr(self, f't_{i}').clear()
            getattr(self, f"monto_{i}").clear()

    def salir(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir ?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()
            
class Ui_Salida(QMainWindow):
    def __init__(self,id_user):
        super(Ui_Salida, self).__init__()
        loadUi("./interfaces/salida.ui", self)
        self.id_user=id_user
        self.actionVolver_al_menu_principal.triggered.connect(self.backmenu)
        self.actionSalir.triggered.connect(self.salir)
        self.btn_buscar.clicked.connect(self.busqueda)
        self.btn_xd.clicked.connect(self.cambiarPage)
        self.btn_print.clicked.connect(self.guardarPDF)
        self.paginaInicia = True
        self.btn_clear.clicked.connect(self.limpiar)
        self.valor_numerico = None
        self.calculo_divisa_thread = CalculoDivisaThread()
        self.calculo_divisa_thread.finished.connect(self.on_calculo_divisa_finished)
        self.calculo_divisa_thread.start()
        self.lineEdit_13.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_13,self.lineEdit_14,self.valor_numerico))
        self.lineEdit_44.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_44,self.lineEdit_43,self.valor_numerico))
        self.lineEdit_47.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_47,self.lineEdit_46,self.valor_numerico))
        self.lineEdit_50.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_50,self.lineEdit_49,self.valor_numerico))
        self.lineEdit_53.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_53,self.lineEdit_52,self.valor_numerico))
        self.lineEdit_56.textChanged.connect(lambda : self.convertir_divisa(self.lineEdit_56,self.lineEdit_55,self.valor_numerico))
        self.total = 0
        self.totalDolar=0
        self.verifytipoUser()

        self.usuario =None
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Cedula FROM Pacientes")
        sugerencias  = [str(row[0]) for row in cursor.fetchall()]
        modelo_completer = QCompleter(sugerencias, self)
        modelo_completer.setCaseSensitivity(0)
        self.in_busqueda.setCompleter(modelo_completer)
        conexion.close()
        
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?", (self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser == "Administrador":
                self.usuario = "Administrador"
            elif tipoUser == "Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontró ningún tipo")
            
                   
    def totalDolares(self):
        total_formateado = "{:,.0f}".format(self.totalDolar).replace(',', '.')
        self.lineEdit_58.setText(total_formateado)
    def totalBs(self):
        total_formateado = "{:,.0f}".format(self.total).replace(',', '.')
        self.lineEdit_57.setText(total_formateado)
    def verificarNumeros(self,cadena):
        patron = r'^\d+$'
        return re.match(patron, cadena) is not None
    
    def convertir_divisa(self, input_line_edit, output_line_edit, valor_numerico):
        cantidad = input_line_edit.text()
        if self.verificarNumeros(cantidad):
            resultado = float(cantidad) * float(valor_numerico)
            resultado_formateado = "{:,.0f}".format(resultado).replace(',', '.')
            output_line_edit.setText(str(resultado_formateado))
            self.total += resultado
            dolar = float(cantidad)
            self.totalDolar += dolar
            self.totalDolares()
            self.totalBs()
        else:
            QMessageBox.information(self, "Error", "Solo números")
            input_line_edit.clear()
            output_line_edit.clear()
   
        
    def on_calculo_divisa_finished(self, resultado):
        # Este método se llama cuando el hilo ha terminado de calcular la divisa
        self.valor_numerico = resultado
        print(f"El valor_numerico al iniciar el programa es: {self.valor_numerico}")

    def calcularDivisas(self,dolar):
        bolivar = self.valor_numerico
        operacion = bolivar * dolar
        return operacion, bolivar
    
    def guardarPDF(self):
        busqueda = self.in_busqueda.text()
        tratamientos = [{self.lineEdit_9.text() : self.lineEdit_14.text(),
                         self.lineEdit_42.text() : self.lineEdit_43.text(),
                         self.lineEdit_45.text() : self.lineEdit_46.text(),
                         self.lineEdit_48.text() : self.lineEdit_49.text(),
                         self.lineEdit_51.text() : self.lineEdit_52.text(),
                         self.lineEdit_54.text(): self.lineEdit_55.text()
                         
                        }
                        ]
        
        cursos= [
             {'fecha': self.dateEdit_2.text(), 'tratamiento': self.lineEdit_59.text()},
            {'fecha': self.dateEdit_8.text(), 'tratamiento': self.lineEdit_60.text()},
            {'fecha' : self.dateEdit_9.text(), 'tratamiento' : self.lineEdit_61.text()},
             {'fecha' : self.dateEdit_10.text(), 'tratamiento' : self.lineEdit_62.text()},
            {'fecha' : self.dateEdit_11.text(), 'tratamiento' : self.lineEdit_63.text()},
            {'fecha' : self.dateEdit_12.text(), 'tratamiento' : self.lineEdit_64.text()}
            
    # Otros diccionarios...
        ]
        total = self.lineEdit_57.text()
        if not busqueda:
            QMessageBox.information(self,"Falta la cedula","Por favor digite la cedula a buscar")
            return
        from interfaces.crearpdf import crear_pdf
          # Abre el diálogo para seleccionar la ubicación de guardado del PDF
        ruta_salida, _ = QFileDialog.getSaveFileName(self, 'Guardar PDF', '', 'Archivos PDF (*.pdf)')
        if not ruta_salida:
            return
            # Crea y guarda el PDF con los datos filtrados
        crear_pdf(ruta_salida=ruta_salida, cedula=busqueda,tratamientos=tratamientos,precioTotal=total,cursoTratamiento=cursos)
        QMessageBox.information(self, "Guardado correctamente", f"Fue guardado en {ruta_salida}")


        
    def cambiarPage(self):
        if self.paginaInicia:
            self.stackedWidget.setCurrentWidget(self.page_2)
            self.paginaInicia = False
        else:
            self.stackedWidget.setCurrentWidget(self.page)
            self.paginaInicia = True
            
    def limpiar(self):
        self.lineEdit_4.clear()
        self.lineEdit_6.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_5.clear()
        self.lineEdit_7.clear()
        self.lineEdit_8.clear()
        self.textEdit.clear()
        self.textEdit_2.clear()
        self.lineEdit_13.clear()
        self.lineEdit_44.clear()
        self.lineEdit_47.clear()
        self.lineEdit_50.clear()
        self.lineEdit_53.clear()
        self.lineEdit_56.clear()
        self.lineEdit_57.clear()
        self.lineEdit_58.clear()
        self.lineEdit_59.clear()
        self.lineEdit_60.clear()
        self.lineEdit_61.clear()
        self.lineEdit_62.clear()
        self.lineEdit_63.clear()
        self.lineEdit_64.clear()
    def busqueda(self):
        busqueda = self.in_busqueda.text()
       
        if not busqueda:
            QMessageBox.information(self,"Falta la cedula","Por favor digite la cedula a buscar")
            return
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Cedula,Telefono,Nombre,Apellido,Direccion,Hipertension,Coagualcion,Diabates,hipertension_Data,diabate_Data,Alergias,Diagnotico FROM Pacientes WHERE Cedula =?",(busqueda,))
       
        resultado = cursor.fetchone()
        
        
        cursor.execute ("SELECT Tratamiento1,Tratamiento2,Tratamiento3,Tratamiento4,Tratamiento5,Tratamiento6,Fecha_Trata FROM PTrata WHERE Cedula =?",(busqueda,))
        resultado_tratamientos = cursor.fetchone()
        if not resultado:
            QMessageBox.warning(self,'Error','La persona no existe en el sistema')
            return
        if resultado:
            self.lineEdit_4.setText(resultado[0])
            self.lineEdit_6.setText(resultado[1])
            self.lineEdit_2.setText(resultado[2])
            self.lineEdit_3.setText(resultado[3])
            self.lineEdit_5.setText(resultado[4])
            Hipertenso = resultado[5]
            if Hipertenso =="No":
                self.btn_no.setChecked(True)
            if Hipertenso =="Si":
                self.btn_si.setChecked(True)
            Coalugacion = resultado[6]
            if Coalugacion =="No":
                self.btn_no_2.setChecked(True)
            if Coalugacion =="Si":
                self.btn_si_2.setChecked(True)
            Diabetes = resultado[7]
            if Diabetes =="No":
                self.btn_no_3.setChecked(True)
            if Diabetes =="Si":
                self.btn_si_3.setChecked(True)
            self.lineEdit_7.setText(resultado[8])
            self.lineEdit_8.setText(resultado[9])
            self.textEdit.setText(resultado[10])
            self.textEdit_2.setText(resultado[11])
            if resultado_tratamientos:
                self.lineEdit_59.setText(resultado_tratamientos[0])
                self.lineEdit_60.setText(resultado_tratamientos[1])
                self.lineEdit_61.setText(resultado_tratamientos[2])
                self.lineEdit_62.setText(resultado_tratamientos[3])
                self.lineEdit_63.setText(resultado_tratamientos[4])
                self.lineEdit_64.setText(resultado_tratamientos[5])
    def salir(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir ?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()   
             
    def backmenu(self):        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)

            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")

            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
            

class MenuPrincipal(QMainWindow):
    def __init__(self, id_user):
        from plyer import notification
        super(MenuPrincipal, self).__init__()
        loadUi("./interfaces/menu.ui", self)

        self.id_user = id_user
        self.setWindowTitle("MenuPrincipal")
        self.showMaximized()
        self.usuario = None
       
        self.notificacion_mostrada = 0
        self.verifytipoUser()
       
        self.setupUi()
        
   
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?",(self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser =="Administrador":
                self.usuario = "Administrador"
            elif tipoUser=="Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontro ningun tipo")

    def setupUi(self):
        self.frame_opciones.hide()
        self.bt_info.clicked.connect(self.informacionView)
        self.bt_menu.clicked.connect(self.toggle_sidebar)
        self.bt_salir.clicked.connect(self.closesesion)
        self.bt_home.clicked.connect(lambda: self.tabWidget.setCurrentWidget(self.principal_tab))
        self.bt_historial.clicked.connect(self.historiaView)
        self.bt_registro.clicked.connect(self.Historyviews)
        self.bt_paciente.clicked.connect(self.PlacasView)
        self.bt_citas.clicked.connect(self.CitasView)
        self.bt_bdd.clicked.connect(self.BddView)
        self.bt_montos.clicked.connect(self.MontosView)
        self.bt_help.clicked.connect(self.ayuda)
        self.bt_act.clicked.connect(self.act_T)
        
        self.bt_buscar.clicked.connect(self.buscar)
        
        self.filtro = self.findChild(QtWidgets.QComboBox, "filtro")
        self.filtro.addItem("Seleccione una opción para filtrar")
        self.filtro.addItems(["Dentista","Fecha_Cita", "Hora_Cita", "Estatus_Cita"])
        self.in_buscar.textChanged.connect(self.buscar)
        self.bt_closesesion.clicked.connect(self.eliminar_datos_acceso)
        
        if self.usuario == "Doctor":
        # En el caso del Doctor, ocultar la columna de usuario y nombre de usuario
            self.tabla_cita.setColumnHidden(0, True) 
            self.tabla_cita.setColumnHidden(1, True) 
        else:
        # Para otros roles, mostrar todas las columnas
            self.tabla_cita.setColumnHidden(0, False)
            self.tabla_cita.setColumnHidden(1, False)
        
        self.act_T()
    def cargardatos(self):
        if self.usuario=="Usuario":
            self.buscar()
        elif self.usuario =="Administrador":
            self.buscar()
        elif self.usuario =="Doctor":
            self.buscar()
    def historiaView(self):
        if self.usuario =="Usuario":
            QMessageBox.information(self,"Permiso Denegado","No tienes permisos para entrar")
            return
        reply = self.showConfirmation("¿Deseas ir al formulario para la creación de las historias?")
        if reply == QMessageBox.Yes:
            historia = Ui_Salida(self.id_user)
            widget.addWidget(historia)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            historia.show()
            self.hide() 
               
    def BddView(self):
        reply = self.showConfirmation("¿Deseas ir al formulario para visualizar todos los pacientes registrados?")
        if reply == QMessageBox.Yes:
            bdd = Ui_pacientes_view(self.id_user)
            widget.addWidget(bdd)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            bdd.show()
            self.hide()  
             
    def MontosView(self):
        if self.usuario =="Usuario":
            QMessageBox.information(self,"Permiso Denegado","No tienes permisos para entrar")
            return
        else:
            reply = self.showConfirmation("¿Deseas ir al formulario para cambiar los montos de los tratamientos?")
            if reply == QMessageBox.Yes:
                montos = Ui_montos(self.id_user)
                widget.addWidget(montos)
                widget.setCurrentIndex(widget.currentIndex() + 1)
                montos.show()
                self.hide()
             
    def act_T(self):
        if self.usuario=="Usuario":
            self.cargarCitasSecretaria()
        elif self.usuario =="Administrador":
            self.cargarCitasSecretaria()
        elif self.usuario =="Doctor": 
            self.cargarCitas()
    def buscar(self):
        filtro = self.filtro.currentText()
        valor = self.in_buscar.text()

        if filtro == "Seleccione una opción para filtrar":
            QtWidgets.QMessageBox.warning(self, "Error", "Debe seleccionar un filtro para buscar")
        elif len(valor) == 0:
            self.act_T()
        elif not valor:
            QtWidgets.QMessageBox.warning(self, "Por favor", "Ingrese alguna especificación de la cita para realizar la búsqueda")
        elif self.tabla_cita.rowCount() == 0:
            QtWidgets.QMessageBox.warning(self, "Advertencia", "No se ha encontrado ningún resultado")
            self.in_buscar.clear()
            self.act_T()
        else:
            # Modificar el filtro "Dentista" para buscar por nombre de dentista
            if filtro == "Dentista":
                filtro = "Nombre_usuario"
            if self.usuario=="Usuario" or self.usuario=="Administrador":
                self.act_T()
                self.cargarCitasSecretaria(filtro,valor)
            else:
                self.act_T()
                self.cargarCitas(filtro, valor)
   


    def cargarCitasSecretaria(self, filtro=None, valor=None):
        self.tabla_cita.setRowCount(0)  # Limpiar la tabla actual
        headers = ["ID del Dentista", "Nombre del Dentista", "Cedula del paciente", "Nombre del paciente", "Apellido del paciente", "Fecha de la cita", "Hora de la cita", "Estatus de la cita"]
        fecha_mas_cercana = None  # Inicializar la variable antes del bloque if
        title = 'Cita más cercana'  # Inicializar la variable antes del bloque if
        message = 'Este es un mensaje de notificación.'  # Inicializar la variable antes del bloque if
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            if filtro and valor:
                # Modificar la consulta para manejar el filtro "Dentista"
                if filtro == "Nombre_usuario":
                    cursor.execute("""
                        SELECT 
                            Users.ID as ID_usuario,
                            Users.Nombres as Nombre_usuario,
                            Pacientes.Cedula, 
                            Pacientes.Nombre, 
                            Pacientes.Apellido, 
                            Cita.Fecha_Cita, 
                            Cita.Hora_Cita, 
                            Cita.Estatus_Cita
                        FROM 
                            Pacientes
                        INNER JOIN 
                            Cita ON Pacientes.Cedula = Cita.Cedula
                        INNER JOIN
                            Users ON Pacientes.ID_user = Users.ID
                        WHERE 
                            {} LIKE ? 
                        ORDER BY 
                            Cita.Fecha_Cita ASC
                    """.format(filtro), ('%' + valor + '%',))
                else:
                    cursor.execute("""
                        SELECT 
                            Users.ID as ID_usuario,
                            Users.Nombres as Nombre_usuario,
                            Pacientes.Cedula, 
                            Pacientes.Nombre, 
                            Pacientes.Apellido, 
                            Cita.Fecha_Cita, 
                            Cita.Hora_Cita, 
                            Cita.Estatus_Cita
                        FROM 
                            Pacientes
                        INNER JOIN 
                            Cita ON Pacientes.Cedula = Cita.Cedula
                        INNER JOIN
                            Users ON Pacientes.ID_user = Users.ID
                        WHERE 
                            {} LIKE ? 
                        ORDER BY 
                            Cita.Fecha_Cita ASC
                    """.format(filtro), ('%' + valor + '%',))
            else:
                cursor.execute("""
                    SELECT 
                        Users.ID as ID_usuario,
                        Users.Nombres as Nombre_usuario,
                        Pacientes.Cedula, 
                        Pacientes.Nombre, 
                        Pacientes.Apellido, 
                        Cita.Fecha_Cita, 
                        Cita.Hora_Cita, 
                        Cita.Estatus_Cita
                    FROM 
                        Pacientes
                    INNER JOIN 
                        Cita ON Pacientes.Cedula = Cita.Cedula
                    INNER JOIN
                        Users ON Pacientes.ID_user = Users.ID
                    ORDER BY 
                        Cita.Fecha_Cita ASC
                """)
            citas = cursor.fetchall() 

            if citas:
                fecha_mas_cercana = datetime.datetime.strptime(citas[0][5], "%Y-%m-%d").date()  # Tomar la fecha de la primera cita

            from plyer import notification
           
            # Mostrar notificación con la fecha de la primera cita
            if self.notificacion_mostrada==0 and fecha_mas_cercana is not None:
                mensaje = f"La cita más cercana es la del paciente {citas[0][3]} {citas[0][4]},\nel dia {citas[0][5]} a las {citas[0][6]}"
                
                # Puedes personalizar los parámetros según tus necesidades
                notification.notify(
                    title=title,
                    message=mensaje,
                    app_name='Consultas Medicas',
                    timeout=10  # Duración en segundos que la notificación estará visible
                )
              
                self.notificacion_mostrada = 1
            elif self.notificacion_mostrada==0:
                # Puedes personalizar los parámetros según tus necesidades
                notification.notify(
                    title=title,
                    message="No hay citas actualmente",
                    app_name='Consultas Medicas',
                    timeout=10  # Duración en segundos que la notificación estará visible
                )
                self.notificacion_mostrada = 1
            self.tabla_cita.setColumnCount(len(headers))
            self.tabla_cita.setHorizontalHeaderLabels(headers)

            for row, cita in enumerate(citas):
                self.tabla_cita.insertRow(row)
                for column, value in enumerate(cita):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.tabla_cita.setItem(row, column, item)

            conexion.close()
                    
        except sqlite3.Error as e:
            QtWidgets.QMessageBox.critical(self, "Error", "Error al consultar la base de datos: " + str(e))
    
    def cargarCitas(self, filtro=None, valor=None):
        self.tabla_cita.setRowCount(0)  # Limpiar la tabla actual
        headers = ["ID del Dentista", "Nombre del Dentista", "Cedula del paciente", "Nombre del paciente", "Apellido del paciente", "Fecha de la cita", "Hora de la cita", "Estatus de la cita"]

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            if filtro and valor:
                # Modificar la consulta para manejar el filtro "Dentista"
                if filtro == "Nombre_usuario":
                    cursor.execute("""
                        SELECT 
                            Users.ID as ID_usuario,
                            Users.Nombres as Nombre_usuario,
                            Pacientes.Cedula, 
                            Pacientes.Nombre, 
                            Pacientes.Apellido, 
                            Cita.Fecha_Cita, 
                            Cita.Hora_Cita, 
                            Cita.Estatus_Cita
                        FROM 
                            Pacientes
                        INNER JOIN 
                            Cita ON Pacientes.Cedula = Cita.Cedula
                        INNER JOIN
                            Users ON Pacientes.ID_user = Users.ID
                        WHERE 
                            {} LIKE ? AND Pacientes.ID_user = ? 
                        ORDER BY 
                            Cita.Fecha_Cita ASC
                        """.format(filtro), ('%' + valor + '%', self.id_user))
                else:
                    cursor.execute("""
                        SELECT 
                            Users.ID as ID_usuario,
                            Users.Nombres as Nombre_usuario,
                            Pacientes.Cedula, 
                            Pacientes.Nombre, 
                            Pacientes.Apellido, 
                            Cita.Fecha_Cita, 
                            Cita.Hora_Cita, 
                            Cita.Estatus_Cita
                        FROM 
                            Pacientes
                        INNER JOIN 
                            Cita ON Pacientes.Cedula = Cita.Cedula
                        INNER JOIN
                            Users ON Pacientes.ID_user = Users.ID
                        WHERE 
                            {} LIKE ? AND Pacientes.ID_user = ? 
                        ORDER BY 
                            Cita.Fecha_Cita ASC
                        """.format(filtro), ('%' + valor + '%', self.id_user))
            else:
                cursor.execute("""
                    SELECT 
                        Users.ID as ID_usuario,
                        Users.Nombres as Nombre_usuario,
                        Pacientes.Cedula, 
                        Pacientes.Nombre, 
                        Pacientes.Apellido, 
                        Cita.Fecha_Cita, 
                        Cita.Hora_Cita, 
                        Cita.Estatus_Cita
                    FROM 
                        Pacientes
                    INNER JOIN 
                        Cita ON Pacientes.Cedula = Cita.Cedula
                    INNER JOIN
                        Users ON Pacientes.ID_user = Users.ID
                    WHERE 
                        Pacientes.ID_user = ? 
                    ORDER BY 
                        Cita.Fecha_Cita ASC
                    """, (self.id_user,))
            citas = cursor.fetchall() 
            fecha_mas_cercana = None

            if citas:
                fecha_mas_cercana = datetime.datetime.strptime(citas[0][5], "%Y-%m-%d").date()  # Tomar la fecha de la primera cita
            from plyer import notification

            # Mostrar notificación con la fecha de la primera cita
            if self.notificacion_mostrada==0 and fecha_mas_cercana is not None:
                mensaje = f"La cita más cercana es la del paciente {citas[0][3]} {citas[0][4]},\nel dia {citas[0][5]} a las {citas[0][6]}"
                
                title = 'Cita má cercana'
                message = 'Este es un mensaje de notificación.'
                # Puedes personalizar los parámetros según tus necesidades
                notification.notify(
                    title=title,
                    message=mensaje,
                    app_name='Consultas Medicas',
                    timeout=10  # Duración en segundos que la notificación estará visible
                )
                self.notificacion_mostrada= 1
            else:
                title = 'Cita má cercana'
                notification.notify(
                    title=title,
                    message="No hay citas actualmente",
                    app_name='Consultas Medicas',
                    timeout=10  # Duración en segundos que la notificación estará visible
                )
                self.notificacion_mostrada= 1
            self.tabla_cita.setColumnCount(len(headers))
            self.tabla_cita.setHorizontalHeaderLabels(headers)

            for row, cita in enumerate(citas):
                self.tabla_cita.insertRow(row)
                for column, value in enumerate(cita):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.tabla_cita.setItem(row, column, item)

            conexion.close()
        except sqlite3.Error as e:
            QtWidgets.QMessageBox.critical(self, "Error", "Error al consultar la base de datos: " + str(e))

    def PlacasView(self):
        if self.usuario =="Usuario":
            QMessageBox.information(self,"Permiso Denegado","No tienes permisos para entrar")
            return
        else:
            reply = self.showConfirmation("¿Deseas ir al formulario de placas?")
            if reply == QMessageBox.Yes:
                placa_view = Ui_placas(self.id_user)
                widget.addWidget(placa_view)
                widget.setCurrentIndex(widget.currentIndex() + 1)
                placa_view.show()
                self.hide()
            
    def ayuda(self):
        dialog = helpView()
        dialog.exec_()
    def CitasView(self):
        reply = self.showConfirmation("¿Deseas ir al formulario de citas?")
        if reply == QMessageBox.Yes:
            citas = Ui_CitasMenu(self.id_user)
            widget.addWidget(citas)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            citas.show()
            self.hide()

    def informacionView(self):
        if self.usuario =="Usuario":
            QMessageBox.information(self,"Permiso Denegado","No tienes permisos para entrar")
            return
        else:
            reply = self.showConfirmation("¿Deseas ir al formulario de cambiar data?")
            if reply == QMessageBox.Yes:
                doctorView = EditDoctor(self.id_user)
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                cursor.execute("SELECT Cedula, Especialidad, Nombres, Apellidos, Sexo, Edad, Direccion, Telefono, Mail, Imagen FROM Users WHERE ID = ?", (self.id_user,))
                resultado = cursor.fetchone()
                if resultado:
                    self.populateDoctorView(doctorView, resultado)
                    widget.addWidget(doctorView)
                    widget.setCurrentIndex(widget.currentIndex() + 1)
                    self.hide()

    def populateDoctorView(self, doctorView, data):
        doctorView.in_cedula_2.setText(data[0])
        doctorView.in_espec_2.setText(data[1])
        doctorView.in_name_2.setText(data[2])
        doctorView.in_apell_2.setText(data[3])
        doctorView.in_age_2.setText(data[5])
        doctorView.in_dir_2.setText(data[6])
        doctorView.in_number_2.setText(data[7])
        doctorView.in_mail_2.setText(data[8])
        if data[4] == "Masculino":
            doctorView.btn_m_2.setChecked(True)
        elif data[4] == "Femenino":
            doctorView.btn_f_2.setChecked(True)
        pixmap1 = QPixmap()
        pixmap1.loadFromData(data[9])
        doctorView.foto_2.setPixmap(pixmap1)

    def Historyviews(self):
        reply = self.showConfirmation("¿Desea ir al formulario de registro de pacientes?")
        if reply == QMessageBox.Yes:
            historia = historiaMenu(self.id_user )
            widget.addWidget(historia)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            historia.show()
            self.hide()

    def showConfirmation(self, message):
        return QMessageBox.question(self, 'Confirmación', message, QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

    def toggle_sidebar(self):
        self.frame_opciones.setHidden(not self.frame_opciones.isHidden())

    def closesesion(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea volver al menu de inicio de sesión?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            login = IngresoUsuario()
            widget.addWidget(login)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            login.show()
            
    def eliminar_datos_acceso(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea eliminar su sesión actual y salir del sistema?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            archivo_datos_acceso = "datos_acceso.json"

            try:
                os.remove(archivo_datos_acceso)
                print(f"Datos de acceso eliminados correctamente.")
                QApplication.quit()
            except FileNotFoundError:
                print("No se encontraron datos de acceso para eliminar.")
                QApplication.quit()
                

class Ui_pacientes_view(QtWidgets.QMainWindow):
    def __init__(self, id_user):
        super(Ui_pacientes_view, self).__init__()
        loadUi("./interfaces/paciente_view.ui", self)
        self.actionVolver_al_menu_principal.triggered.connect(self.back_menu)
        self.actionSalir.triggered.connect(self.salir)
        self.bt_act.clicked.connect(self.act_T)
        self.bt_buscar.clicked.connect(self.buscar)
        self.id_user = id_user
        self.usuario = None  # Agregamos una variable de instancia para almacenar el tipo de usuario
        self.verifytipoUser()  # Llamamos a la función de verificación de usuario
        self.act_T()
        self.filtro = self.findChild(QtWidgets.QComboBox, "filtro")
        self.filtro.addItem("Seleccione una opción para filtrar")
        self.filtro.addItems(["Dentista", "Cedula", "Nombre", "Edad", "Sexo", "Direccion", "Fecha_Diagnotico"])
        self.in_buscar.textChanged.connect(self.buscar)
       
        self.bt_preview.clicked.connect(self.guardarPDF)
        self.bt_backup.clicked.connect(self.exportar_a_excel)
        self.bt_import.clicked.connect(self.importarData)
        
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?", (self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser == "Administrador":
                self.usuario = "Administrador"
            elif tipoUser == "Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontró ningún tipo")
    
    def importarData(self):
        # Conectar a la base de datos SQLite
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()

        # Abrir el archivo Excel
        archivo_excel, _ = QFileDialog.getOpenFileName(self, 'Seleccionar archivo Excel', '', 'Archivos Excel (*.xlsx)')
        if not archivo_excel:
            return  # Si no se selecciona ningún archivo, salir

        # Cargar el archivo Excel
        libro_excel = openpyxl.load_workbook(archivo_excel)

        # Leer datos de la hoja 'Pacientes'
        hoja_pacientes = libro_excel['Pacientes']
        datos_pacientes = []
        for fila in hoja_pacientes.iter_rows(min_row=2, values_only=True):
            datos_pacientes.append(fila)

        # Insertar datos en la tabla Pacientes, evitando duplicados
        for dato in datos_pacientes:
                cedula = dato[1]  # La columna Cedula es el segundo elemento en los datos
                cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula=?", (cedula,))
                if cursor.fetchone()[0] == 0:
                    cursor.execute("""
                        INSERT INTO Pacientes (ID_user, Cedula, Nombre, Apellido, Edad, Direccion, Sexo, Fecha_Diagnotico)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, dato)

        # Leer datos de la hoja 'Cita'
        hoja_cita = libro_excel['Cita']
        datos_cita = []
        for fila in hoja_cita.iter_rows(min_row=2, values_only=True):
            datos_cita.append(fila)

        # Insertar datos en la tabla Cita, evitando duplicados
        for dato in datos_cita:
            id_cita = dato[2]  # La columna ID_Cita es el tercer elemento en los datos
            cursor.execute("SELECT COUNT(*) FROM Cita WHERE ID_Cita=?", (id_cita,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    INSERT INTO Cita (ID_user, Cedula, ID_Cita, Fecha_Cita, Hora_Cita, Estatus_Cita)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, dato)

        # Leer datos de la hoja 'Trata'
        hoja_trata = libro_excel['Trata']
        datos_trata = []
        for fila in hoja_trata.iter_rows(min_row=2, values_only=True):
            datos_trata.append(fila)

        # Insertar datos en la tabla Trata, evitando duplicados
        for dato in datos_trata:
            num_trata = dato[0]  # La columna Num_trata es el primer elemento en los datos
            cursor.execute("SELECT COUNT(*) FROM Trata WHERE Num_trata=?", (num_trata,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    INSERT INTO Trata (Num_trata, tipo_tratamiento, tratamiento, monto)
                    VALUES (?, ?, ?, ?)
                """, dato)

        # Leer datos de la hoja 'Ptrata'
        hoja_ptrata = libro_excel['Ptrata']
        datos_ptrata = []
        for fila in hoja_ptrata.iter_rows(min_row=2, values_only=True):
            datos_ptrata.append(fila)

        # Insertar datos en la tabla Ptrata, evitando duplicados
        for dato in datos_ptrata:
            id_trata = dato[0]  # La columna ID_Trata es el primer elemento en los datos
            cursor.execute("SELECT COUNT(*) FROM Ptrata WHERE ID_Trata=?", (id_trata,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    INSERT INTO Ptrata (ID_Trata, Cedula, Tipo_Trata1, Tratamiento1, Tipo_Trata2, Tratamiento2,
                    Tipo_Trata3, Tratamiento3, Tipo_Trata4, Tratamiento4, Tipo_Trata5, Tratamiento5, Tipo_Trata6, 
                    Tratamiento6, Fecha_Trata)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, dato)

        # Guardar cambios en la base de datos
        conexion.commit()

        # Cerrar la conexión
        conexion.close()
        QMessageBox.information(self, "Información", "Se han guardado correctamente los datos importados.")


           
    def exportar_a_excel(self):
        # Conectar a la base de datos SQLite
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()

        # Crear un nuevo archivo Excel
        libro_excel = openpyxl.Workbook()

        # Obtener datos de la tabla Pacientes
        cursor.execute("""
            SELECT 
                ID_user,
                Cedula, 
                Nombre, 
                Apellido, 
                Edad, 
                Direccion, 
                Sexo, 
                Fecha_Diagnotico
            FROM 
                Pacientes
            ORDER BY 
                Fecha_Diagnotico ASC
        """)
        datos_pacientes = cursor.fetchall()
        columnas_pacientes = [description[0] for description in cursor.description]

        # Escribir datos de la tabla Pacientes en la hoja 'Pacientes'
        hoja_pacientes = libro_excel.create_sheet('Pacientes')
        hoja_pacientes.append(columnas_pacientes)
        for dato in datos_pacientes:
            hoja_pacientes.append(list(dato))

        # Obtener datos de la tabla Cita
        cursor.execute("""
            SELECT * FROM Cita 
        """)
        datos_cita = cursor.fetchall()
        columnas_cita = [description[0] for description in cursor.description]

        # Escribir datos de la tabla Cita en la hoja 'Cita'
        hoja_cita = libro_excel.create_sheet('Cita')
        hoja_cita.append(columnas_cita)
        for dato in datos_cita:
            hoja_cita.append(list(dato))

        # Obtener datos de la tabla Trata
        cursor.execute("""
            SELECT * FROM Trata
        """)
        datos_trata = cursor.fetchall()
        columnas_trata = [description[0] for description in cursor.description]

        # Escribir datos de la tabla Trata en la hoja 'Trata'
        hoja_trata = libro_excel.create_sheet('Trata')
        hoja_trata.append(columnas_trata)
        for dato in datos_trata:
            hoja_trata.append(list(dato))

        # Obtener datos de la tabla Ptrata
        cursor.execute("""
            SELECT * FROM PTrata
        """)
        datos_ptrata = cursor.fetchall()
        columnas_ptrata = [description[0] for description in cursor.description]

        # Escribir datos de la tabla Ptrata en la hoja 'Ptrata'
        hoja_ptrata = libro_excel.create_sheet('Ptrata')
        hoja_ptrata.append(columnas_ptrata)
        for dato in datos_ptrata:
            hoja_ptrata.append(list(dato))

        # Guardar el archivo Excel con el filtro específico
        filtro = 'Archivos Excel (*.xlsx);;Todos los archivos (*)'
        archivo_excel, _ = QFileDialog.getSaveFileName(self, 'Guardar como', '', filtro)

        if archivo_excel:
            if not archivo_excel.endswith('.xlsx'):
                archivo_excel += '.xlsx'

            libro_excel.save(archivo_excel)
    
        QMessageBox.information(self, "Información", "Se ha guardado correctamente.")

        conexion.close()
        
    def guardarPDF(self):
        from interfaces.pdfReportegeneral import recuperardatos_Doctor,recuperar_datos_bd,crear_pdf
        try:
            # Selecciona la función de recuperación de datos según el tipo de usuario
            if self.usuario == "Usuario" or self.usuario == "Administrador":
                datos = recuperar_datos_bd()
            elif self.usuario == "Doctor":
                datos = recuperardatos_Doctor(self.id_user)
            else:
                datos = None

            if not datos:
                QMessageBox.warning(self, "Advertencia", "No hay datos para generar el informe.")
                return

            # Filtra los datos según la opción seleccionada en el combo de filtro
            filtro = self.filtro.currentText()
            valor = self.in_buscar.text()

            if filtro != "Seleccione una opción para filtrar" and valor:
                # Filtra los datos según la opción seleccionada y el valor ingresado
                datos = [dato for dato in datos if valor.lower() in str(dato).lower()]

            # Abre el diálogo para seleccionar la ubicación de guardado del PDF
            ruta_salida, _ = QFileDialog.getSaveFileName(self, 'Guardar PDF', '', 'Archivos PDF (*.pdf)')

            if not ruta_salida:
                return

            # Crea y guarda el PDF con los datos filtrados
            crear_pdf(ruta_salida=ruta_salida, datos=datos)

            QMessageBox.information(self, "Guardado correctamente", f"Fue guardado en {ruta_salida}")

        except Exception as e:
            print(f"Error en guardarPDF: {e}")
                   
    def back_menu(self):
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)
          
            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")
            
            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
    
    def act_T(self):
        if self.usuario == "Usuario":
            self.cargarCitasSecretaria()
        elif self.usuario == "Administrador":
            self.cargarCitasSecretaria()
        elif self.usuario == "Doctor":
            self.cargarCitas()

    
    def cargarCitasSecretaria(self,filtro=None,valor=None):
        self.tabla_p.setRowCount(0)  # Limpiar la tabla actual
        headers = ["ID del usuario", "Nombre del usuario", "Cedula del paciente", "Nombre del paciente", "Apellido del paciente", "Edad", "Dirección", "Sexo", "Fecha del diagnóstico"]
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            if filtro and valor:
                # Modificar la consulta para manejar el filtro "Dentista"
                if filtro == "Nombre_usuario":
                   cursor.execute("""
                         SELECT 
                Users.ID as ID_usuario,
                Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
                Pacientes.Cedula, 
                Pacientes.Nombre, 
                Pacientes.Apellido, 
                Pacientes.Edad, 
                Pacientes.Direccion, 
                Pacientes.Sexo, 
                Pacientes.Fecha_Diagnotico
            FROM 
                Pacientes
            INNER JOIN
                Users ON Pacientes.ID_user = Users.ID
            WHERE 
                Users.Nombres || ' ' || Users.Apellidos LIKE ?
            ORDER BY 
                Pacientes.Fecha_Diagnotico ASC
                        """.format(filtro), ('%' + valor + '%',))
               
                else:
                    cursor.execute("""
                       SELECT 
                Users.ID as ID_usuario,
                Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
                Pacientes.Cedula, 
                Pacientes.Nombre, 
                Pacientes.Apellido, 
                Pacientes.Edad, 
                Pacientes.Direccion, 
                Pacientes.Sexo, 
                Pacientes.Fecha_Diagnotico
            FROM 
                Pacientes
            INNER JOIN
                Users ON Pacientes.ID_user = Users.ID
            WHERE 
                Pacientes.{} LIKE ?
            ORDER BY 
                Pacientes.Fecha_Diagnotico ASC
                        """.format(filtro), ('%' + valor + '%',))
            else:
                cursor.execute("""
                     SELECT 
            Users.ID as ID_usuario,
            Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
            Pacientes.Cedula, 
            Pacientes.Nombre, 
            Pacientes.Apellido, 
            Pacientes.Edad, 
            Pacientes.Direccion, 
            Pacientes.Sexo, 
            Pacientes.Fecha_Diagnotico
        FROM 
            Pacientes
       
        INNER JOIN
            Users ON Pacientes.ID_user = Users.ID
        
        ORDER BY 
            Pacientes.Fecha_Diagnotico ASC
                    """)
            citas = cursor.fetchall() 

            self.tabla_p.setColumnCount(len(headers))
            self.tabla_p.setHorizontalHeaderLabels(headers)

            for row, cita in enumerate(citas):
                self.tabla_p.insertRow(row)
                for column, value in enumerate(cita):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.tabla_p.setItem(row, column, item)

            conexion.close()
                    
        except sqlite3.Error as e:
            QtWidgets.QMessageBox.critical(self, "Error", "Error al consultar la base de datos: " + str(e))

    def cargarCitas(self, filtro=None, valor=None):
        self.tabla_p.setRowCount(0)  # Limpiar la tabla actual
        headers = ["ID del usuario", "Nombre del usuario", "Cedula del paciente", "Nombre del paciente", "Apellido del paciente", "Edad", "Dirección", "Sexo", "Fecha del diagnóstico"]

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            if filtro and valor:
                # Modificar la consulta para manejar el filtro "Dentista"
                if filtro == "Nombre_usuario":
                    cursor.execute("""
                        SELECT 
                Users.ID as ID_usuario,
                Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
                Pacientes.Cedula, 
                Pacientes.Nombre, 
                Pacientes.Apellido, 
                Pacientes.Edad, 
                Pacientes.Direccion, 
                Pacientes.Sexo, 
                Pacientes.Fecha_Diagnotico
            FROM 
                Pacientes
            
            INNER JOIN
                Users ON Pacientes.ID_user = Users.ID
            WHERE 
                Users.Nombres || ' ' || Users.Apellidos LIKE ? AND Pacientes.ID_user = ? 
            ORDER BY 
                Pacientes.Fecha_Diagnotico ASC
                        """.format(filtro), ('%' + valor + '%', self.id_user))
                else:
                    cursor.execute("""
                         SELECT 
                Users.ID as ID_usuario,
                Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
                Pacientes.Cedula, 
                Pacientes.Nombre, 
                Pacientes.Apellido, 
                Pacientes.Edad, 
                Pacientes.Direccion, 
                Pacientes.Sexo, 
                Pacientes.Fecha_Diagnotico
            FROM 
                Pacientes
           
            INNER JOIN
                Users ON Pacientes.ID_user = Users.ID
            WHERE 
                Pacientes.{} LIKE ? AND Pacientes.ID_user = ? 
            ORDER BY 
                Pacientes.Fecha_Diagnotico ASC
                        """.format(filtro), ('%' + valor + '%', self.id_user))
            else:
                cursor.execute("""
                      SELECT 
            Users.ID as ID_usuario,
            Users.Nombres || ' ' || Users.Apellidos as Nombre_completo,
            Pacientes.Cedula, 
            Pacientes.Nombre, 
            Pacientes.Apellido, 
            Pacientes.Edad, 
            Pacientes.Direccion, 
            Pacientes.Sexo, 
            Pacientes.Fecha_Diagnotico
        FROM 
            Pacientes
        
        INNER JOIN
            Users ON Pacientes.ID_user = Users.ID
        WHERE 
            Pacientes.ID_user = ? 
        ORDER BY 
            Pacientes.Fecha_Diagnotico ASC
                    """, (self.id_user,))
            citas = cursor.fetchall() 

            self.tabla_p.setColumnCount(len(headers))
            self.tabla_p.setHorizontalHeaderLabels(headers)

            for row, cita in enumerate(citas):
                self.tabla_p.insertRow(row)
                for column, value in enumerate(cita):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    self.tabla_p.setItem(row, column, item)

            conexion.close()
        except sqlite3.Error as e:
            QtWidgets.QMessageBox.critical(self, "Error", "Error al consultar la base de datos: " + str(e))


    
    def buscar(self):
        filtro = self.filtro.currentText()
        valor = self.in_buscar.text()
        if filtro == "Seleccione una opción para filtrar":
            QtWidgets.QMessageBox.warning(self, "Error", "Debe seleccionar un filtro para buscar")
        elif len(valor) == 0:
            self.act_T()
        elif not valor:
            QtWidgets.QMessageBox.warning(self, "Por favor", "Ingrese alguna especificación del paciente para realizar la búsqueda")
        elif self.tabla_p.rowCount() == 0:
            QtWidgets.QMessageBox.warning(self, "Advertencia", "No se ha encontrado ningún registro")  
        else:
            # Modificar el filtro "Dentista" para buscar por nombre de dentista
            if filtro == "Dentista":
                filtro = "Nombre_usuario"
            
            if self.usuario == "Usuario" or self.usuario == "Administrador":  # Corregir aquí
                self.cargarCitasSecretaria(filtro, valor)
            elif self.usuario == "Doctor":
                self.cargarCitas(filtro, valor)

    def salir(self):
        reply = QtWidgets.QMessageBox.question(
            self, 'Confirmación', '¿Desea Salir?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes
        )
        if reply == QtWidgets.QMessageBox.Yes:
            QtWidgets.QApplication.quit()

class helpView(QDialog):
    def __init__(self):
        super(helpView,self).__init__()
        loadUi('interfaces/help.ui', self)
        self.setWindowTitle('Información')
       
        self.click.setText('<a href="https://github.com/Reinaldito04/App-Consultas-y-Historias-Medicas">Haz Click aqui</a>')
        self.click.setOpenExternalLinks(True)
        self.click.linkActivated.connect(self.open_github_link)
        self.click.setStyleSheet("QLabel { text-decoration: none; }")

    def open_github_link(self):
        QDesktopServices.openUrl(QUrl('https://github.com/Reinaldito04/App-Consultas-y-Historias-Medicas'))

class EditDoctor(QMainWindow):
    def __init__(self,id_user):
        super(EditDoctor, self).__init__()
        self.id_user = id_user
        loadUi("interfaces/edicion.ui", self)
        self.btn_passwordChange.clicked.connect(self.PasswordView)
        self.bt_delete.clicked.connect(self.eliminarInfo)
        self.btn_save.clicked.connect(self.modifyInfo)
        self.actionBack.triggered.connect(self.back_menu)
        self.btn_adduser.clicked.connect(self.ingresoRegistro)
        
    def ingresoRegistro(self):
        reply = QtWidgets.QMessageBox.question(
            self, 'Confirmación', '¿Desea ir al formulario para añadir usuarios al sistema?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes
        )
        if reply == QtWidgets.QMessageBox.Yes:
            registroview = Registro(self.id_user)
            widget.addWidget(registroview)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            registroview.show()
            self.hide()
        
    def back_menu(self):
        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)
          
            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")
            
            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
            
    def modifyInfo(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            'Deseas cambiar tu información personal?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
            
        if self.btn_m_2.isChecked():
               sexo = "Masculino"
        if self.btn_f_2.isChecked():
                sexo = "Femenino"
        if reply == QMessageBox.Yes:
           conexion = sqlite3.connect('interfaces/database.db')
           cursor = conexion.cursor()
           cedula = self.in_cedula_2.text()
           especialidad = self.in_espec_2.text()
           nombre = self.in_name_2.text()
           apellido = self.in_apell_2.text()
           direccion = self.in_dir_2.text()
           telefono  = self.in_number_2.text()
           edad = self.in_age_2.text()
           mail = self.in_mail_2.text()
           
           
           cursor.execute("UPDATE Users SET Cedula = ?, Especialidad = ?, Nombres = ?, Apellidos = ?, Direccion = ?, Telefono = ?, Mail = ?, Sexo = ?, Edad = ? WHERE ID = ?", (cedula, especialidad, nombre, apellido, direccion, telefono, mail, sexo, edad, self.id_user))
           QMessageBox.information(self,"Realizado","Los cambios han sido guardados correctamente")
           conexion.commit()    
           conexion.close()
           
    def eliminarInfo(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Deseas ir al formulario para eliminar su usuario?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply ==QMessageBox.Yes:
            dialog = DeleteAllData(self.id_user)
            dialog.exec_()
            
            
    def PasswordView(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Deseas cambiar tu contraseña?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply ==QMessageBox.Yes:
            dialog = PasswordMenu(self.id_user)
            dialog.exec_()
            
class DeleteAllData(QDialog):
    def __init__(self,id_user):
        super(DeleteAllData , self).__init__()
        self.id_user = id_user
        loadUi("interfaces/eliminarData.ui", self)
        self.bt_back.clicked.connect(self.back)
        self.bt_delete.clicked.connect(self.deleteData)
        self.setWindowTitle("Eliminar Datos")

    def cifrar_contrasenia(self, contrasenia):
        # Cifrar la contraseña usando un algoritmo de hash (SHA-256 en este caso)
        cifrado = hashlib.sha256()
        cifrado.update(contrasenia.encode('utf-8'))
        return cifrado.hexdigest()
    def deleteData(self):
        contraseniaaAntigua = self.ln_password.text()
        contraseniaRepeat = self.ln_repeatPassword.text()
        iduser=self.id_user
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Estas seguro de eliminar tu información?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes :
            try:
                cursor.execute("SELECT Password, Username From Users WHERE ID =? ", (iduser,))
                resultado = cursor.fetchone()
                if resultado and resultado[0] ==  self.cifrar_contrasenia(contraseniaaAntigua):
                    if contraseniaaAntigua  == contraseniaRepeat :
                       
                        nombre = resultado[1]
                        cursor.execute('BEGIN TRANSACTION;')
                        cursor.execute("DELETE FROM Users WHERE ID =?",(self.id_user,))
                        cursor.execute('DELETE FROM Pacientes WHERE ID_user = ?', (self.id_user,))
                        conexion.commit()
                        QMessageBox.information(self,'Exito',f'Felicidades {nombre} tus datos han sido eliminado correctamente')
                        registro = Registro()
                        widget.addWidget(registro)
                        widget.setCurrentIndex(widget.currentIndex()+1)
                        
                    else:
                        QMessageBox.information(self,'Error','Las contraseñas no coinciden.')
                else:
                    QMessageBox.information(self,'Error','La contraseña es incorrecta.')
            except sqlite3.Error as e:
                print(f"Error de base de datos: {e}")
            finally:
                conexion.close()
         
    def back(self):
            doctorView = EditDoctor(self.id_user)
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            cursor.execute("SELECT Cedula,Especialidad,Nombres,Apellidos,Sexo,Edad,Direccion,Telefono,Mail , Imagen FROM Users WHERE ID = ?", (self.id_user,))
            resultado = cursor.fetchone()
            if resultado :
                doctorView.in_cedula_2.setText(resultado[0])
                doctorView.in_espec_2.setText(resultado[1])
                doctorView.in_name_2.setText(resultado[2])
                doctorView.in_apell_2.setText(resultado[3])
                doctorView.in_age_2.setText(resultado[5])
                doctorView.in_dir_2.setText(resultado[6])
                doctorView.in_number_2.setText(resultado[7])
                doctorView.in_mail_2.setText(resultado[8])
                sexo = resultado[4]
                if sexo == "Masculino":
                    doctorView.btn_m_2.setChecked(True)
                if sexo == "Femenino":
                    doctorView.btn_f_2.setChecked(True)
                pixmap1 = QPixmap()
                pixmap1.loadFromData(resultado[9])
                doctorView.foto_2.setPixmap(pixmap1)
            widget.addWidget(doctorView)
            widget.setCurrentIndex(widget.currentIndex()+1)
            
            self.hide()
        
class Ui_CitasMenu(QMainWindow):
    def __init__(self,id_user):
        super(Ui_CitasMenu, self).__init__()
        self.id_user = id_user
        loadUi("interfaces/citas.ui", self)
        self.actionVolver_al_menu_principal.triggered.connect(self.back)
        self.actionSalir.triggered.connect(self.salir)
        self.btn_buscar.clicked.connect(self.searchdata)
        self.btn_agg.clicked.connect(self.aggCite)
        self.btn_clear.clicked.connect(self.clear)
        self.btn_edit.clicked.connect(self.editarCita)
        #self.btn_delete.clicked.connect(self.eliminarCita)
        self.setWindowTitle("Menu de Citas")
        self.showMaximized()
        self.usuario =None
        self.verifytipoUser()
    
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?",(self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser =="Administrador":
                self.usuario = "Administrador"
            elif tipoUser=="Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontro ningun tipo")
                
    def eliminarCita(self):
        try:
            cedula = self.in_busqueda.text()
    
            if len(cedula) == 0:
                QMessageBox.critical(self, "Error", "Ingrese una cédula")
            else:
              
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                cursor.execute("DELETE FROM Cita WHERE Cedula = ?", (cedula,))

                conexion.commit()
                conexion.close()
        
        # Eliminación exitosa, muestra un mensaje y realiza otras acciones si es necesario
                QMessageBox.information(self, "Realizado", "La cita ha sido eliminada correctamente")
                
                self.clear()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", "Error al eliminar la cita  de la base de datos: " + str(e))
       
    def clear(self):
        self.in_busqueda.clear()
        self.txt_name.clear()
        self.txt_apell.clear()
        self.tableWidget.setRowCount(0)
        
    def editarCita(self):
        idUser = self.id_user
        cedula = self.in_busqueda.text()
        
        fecha = self.fecha.selectedDate()
        fechaToString = fecha.toString('yyyy-MM-dd')
        statusCita = None
        if self.bt_act.isChecked():
            statusCita = 'Activa'
        if self.bt_cancel.isChecked():
            statusCita = 'Cancelada'
        hora = self.hora.time()
        horaToString = hora.toString('hh:mm:ss')
        
        conexion = None  # Inicializa la variable de conexión fuera del bloque try

        try:
            if not cedula:
                QMessageBox.warning(self, "Introduzca una cedula", "Debes introducir una cedula antes de editar")
                return
            reply = QMessageBox.question(
                self,
                'Confirmación',
                '¿Desea cambiar la fecha y hora de la cita?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply == QMessageBox.Yes:
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                
                cursor.execute("SELECT COUNT(*) FROM Cita WHERE Fecha_Cita = ? AND Hora_Cita = ? AND Cedula != ?", (fechaToString, horaToString, cedula))
                existe_cita = cursor.fetchone()[0] > 0

                if existe_cita:
                    QMessageBox.information(self, "Cita", "Ya existe una cita programada para la misma fecha y hora.")
                    return

                # Si el paciente existe y no hay conflicto de citas, proceder a actualizar la cita
                cursor.execute("""
                    UPDATE Cita 
                    SET Fecha_Cita = ?, Hora_Cita = ?, Estatus_Cita = ?
                    WHERE Cedula = ? 
                """, (fechaToString, horaToString, statusCita, cedula))

                # Guardar los cambios en la base de datos
                conexion.commit()

                # Mostrar un mensaje de éxito
                QMessageBox.information(self, "Información", "Cita actualizada con éxito.")
            else:
                # Si el usuario no confirma el cambio, puedes agregar código aquí si es necesario
                pass

        except sqlite3.Error as error:
            # En caso de error, mostrar un mensaje de error
            QMessageBox.critical(self, "Error", f"Error al actualizar la cita: {str(error)}")
        
        finally:
            # Cerrar la conexión con la base de datos en el bloque finally
            if conexion:
                conexion.close()


    def aggCite(self):
        cedula = self.in_busqueda.text()
        
        fecha = self.fecha.selectedDate()
        fechaToString = fecha.toString('yyyy-MM-dd')
        statusCita = None
        if self.bt_act.isChecked():
            statusCita = 'Activa'
        if self.bt_cancel.isChecked():
            statusCita = 'Cancelada'
        hora = self.hora.time()
        horaToString = hora.toString('hh:mm:ss')
        
        try:
            if not cedula:
                QMessageBox.warning(self, "Introduzca una cedula", "Debes introducir una cedula antes de guardar")
                return
            reply = QMessageBox.question(
                self,
                'Confirmación',
                '¿Deseas agregar o editar la cita?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            if reply == QMessageBox.Yes:
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                
                # Verificar si existe un paciente con la cédula proporcionada
                cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula = ?", (cedula,))
                existe_paciente = cursor.fetchone()[0] > 0
                
                if existe_paciente:
                    # Verificar si el paciente ya tiene una cita registrada
                    cursor.execute("SELECT COUNT(*) FROM Cita WHERE Cedula = ?", (cedula,))
                    tiene_cita = cursor.fetchone()[0] > 0
                    
                    if tiene_cita:
                        # Si el paciente ya tiene una cita, mostrar un mensaje informativo y sugerir editar la cita existente
                        QMessageBox.information(self, "Información", "El paciente ya tiene una cita registrada. Por favor, edite la cita existente.")
                    else:
                        # Verificar si ya hay una cita programada para la misma fecha y hora y otro paciente
                        cursor.execute("SELECT COUNT(*) FROM Cita WHERE Fecha_Cita = ? AND Hora_Cita = ? AND Cedula != ?", (fechaToString, horaToString, cedula))
                        existe_cita = cursor.fetchone()[0] > 0
                        
                        if existe_cita:
                            QMessageBox.information(self, "Cita", "Ya existe una cita programada para la misma fecha y hora.")
                        else:
                            # Si el paciente no tiene una cita y no hay conflicto de citas, proceder a insertar una nueva cita
                            cursor.execute("""
                                INSERT INTO Cita (Cedula, Fecha_Cita, Hora_Cita, Estatus_Cita, ID_user)
                                VALUES (?, ?, ?, ?, ?)
                            """, (cedula, fechaToString, horaToString, statusCita, self.id_user))

                            # Guardar los cambios en la base de datos
                            conexion.commit()
                            
                            # Mostrar un mensaje de éxito
                            QMessageBox.information(self, "Información", "Cita agregada con éxito.")
                else:
                    # Si el paciente no está registrado, mostrar un mensaje de error
                    QMessageBox.warning(self, "Advertencia", "No se encontró un paciente con la cédula proporcionada.")
                
                # Cerrar la conexión con la base de datos
                conexion.close()

        except sqlite3.Error as error:
            # En caso de error, mostrar un mensaje de error
            QMessageBox.critical(self, "Error", f"Error al agregar o editar la cita: {str(error)}")

    def searchdata(self):
        try:
            busqueda = self.in_busqueda.text()
            idUser = self.id_user

            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            if self.usuario == "Administrador" or self.usuario == "Usuario":
                # Si es un administrador o usuario, puede buscar cualquier cita
                cursor.execute("""
                    SELECT 
                        Pacientes.Cedula, 
                        Pacientes.Nombre, 
                        Pacientes.Apellido, 
                        Cita.Fecha_Cita, 
                        Cita.Hora_Cita, 
                        Cita.Estatus_Cita
                    FROM 
                        Pacientes
                    INNER JOIN 
                        Cita ON Pacientes.Cedula = Cita.Cedula
                    WHERE 
                        Pacientes.Cedula = ?
                """, (busqueda,))
            elif self.usuario == "Doctor":
                # Si es un doctor, solo puede buscar citas asociadas a su ID de usuario
                cursor.execute("""
                    SELECT 
                        Pacientes.Cedula, 
                        Pacientes.Nombre, 
                        Pacientes.Apellido, 
                        Cita.Fecha_Cita, 
                        Cita.Hora_Cita, 
                        Cita.Estatus_Cita
                    FROM 
                        Pacientes
                    INNER JOIN 
                        Cita ON Pacientes.Cedula = Cita.Cedula
                    WHERE 
                        Pacientes.Cedula = ? AND Pacientes.ID_user = ?
                """, (busqueda, idUser))
            else:
                # Manejar el caso en el que el tipo de usuario no sea reconocido
                QMessageBox.critical(self, "Error", "Rol de usuario no reconocido")
                return

            tabla_cita = cursor.fetchall()

            if tabla_cita:
                for cita in tabla_cita:
                    cedula = cita[0]
                    nombre = cita[1]
                    apellido = cita[2]
                    fecha = cita[3]
                    hora = cita[4]
                    estatus = cita[5]

                    # Limpiar la tabla existente si es necesario
                    self.tableWidget.clearContents()

                    # Establecer el número de filas y columnas en la tabla
                    self.tableWidget.setRowCount(len(tabla_cita))
                    self.tableWidget.setColumnCount(len(tabla_cita[0]))

                    # Agregar los datos a la tabla
                    for row, paciente in enumerate(tabla_cita):
                        for column, value in enumerate(paciente):
                            item = QTableWidgetItem(str(value))
                            self.tableWidget.setItem(row, column, item)

                    self.txt_name.setText(nombre)
                    self.txt_apell.setText(apellido)
                    fecha_cita = fecha

                    self.fecha.setSelectedDate(QDate.fromString(fecha_cita, 'yyyy-MM-dd'))
                    hora_cita = QTime.fromString(hora, 'hh:mm:ss')

                    if estatus == 'Activa':
                        self.bt_act.setChecked(True)
                    if estatus == 'Cancelada':
                        self.bt_cancel.setChecked(True)

                    self.hora.setTime(hora_cita)

            else:
                # Manejar el caso en el que no se encontraron resultados
                QMessageBox.warning(self, "Sin resultados", "No se encontraron citas para el paciente")

                # Si deseas mostrar el nombre y apellido incluso si no hay cita, puedes hacerlo aquí
                cursor.execute("SELECT Nombre, Apellido FROM Pacientes WHERE Cedula = ? AND ID_user = ?", (busqueda, idUser))
                datos_paciente = cursor.fetchone()

                if datos_paciente:
                    nombre_paciente, apellido_paciente = datos_paciente
                    self.txt_name.setText(nombre_paciente)
                    self.txt_apell.setText(apellido_paciente)
                else:
                    QMessageBox.warning(self, "Sin resultados", "No se encontró al paciente")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al buscar paciente: {str(error)}")
        finally:
            conexion.close()


    def back(self):
        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)

            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")

            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
            
    def salir(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()
            
class PasswordMenu(QDialog):
    def __init__(self,id_user ):
        super(PasswordMenu, self).__init__()
        self.id_user = id_user
        loadUi("interfaces/password.ui", self)
        self.bt_menu.clicked.connect(self.returnMenu)
        self.bt_passwordChange.clicked.connect(self.cambiarPassword)
        self.setWindowTitle("Cambiar Contraseña")

    def cifrar_contrasenia(self, contrasenia):
        # Cifrar la contraseña usando un algoritmo de hash (SHA-256 en este caso)
        cifrado = hashlib.sha256()
        cifrado.update(contrasenia.encode('utf-8'))
        return cifrado.hexdigest()
    
    def cambiarPassword(self):
        contraseniaaAntigua = self.txt_passwordOld.text()
        contraseniaNew = self.txt_passwordNew.text()
        contraseniaRepeat = self.txt_passowrdRepeat.text()
        iduser=self.id_user
        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Estas seguro de cambiar tu contraseña?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes :
            try:
                cursor.execute("SELECT Password , Username From Users WHERE ID =? ", (iduser,))
                resultado = cursor.fetchone()
                if resultado and resultado[0] ==  self.cifrar_contrasenia(contraseniaaAntigua):
                    if contraseniaNew  == contraseniaRepeat :
                        contraseniaCifrada = self.cifrar_contrasenia(contraseniaNew)
                        nombre = resultado[1]
                        cursor.execute("UPDATE Users SET Password= ? WHERE ID= ?", (contraseniaCifrada , iduser))
                        conexion.commit()
                        QMessageBox.information(self,'Exito',f'Felicidades {nombre} tu contraseña ha sido cambiada correctamente') 
                    else:
                        QMessageBox.information(self,'Error','Las contraseñas no coinciden.')
                else:
                    QMessageBox.information(self,'Error','La contraseña antigua no es valida.')       
            except sqlite3.Error as e:
                print(f"Error de base de datos: {e}")
            finally:
                conexion.close()
    def returnMenu(self):
            doctorView = EditDoctor(self.id_user)
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            cursor.execute("SELECT Cedula,Especialidad,Nombres,Apellidos,Sexo,Edad,Direccion,Telefono,Mail , Imagen FROM Users WHERE ID = ?", (self.id_user,))
            resultado = cursor.fetchone()
            if resultado :
                doctorView.in_cedula_2.setText(resultado[0])
                doctorView.in_espec_2.setText(resultado[1])
                doctorView.in_name_2.setText(resultado[2])
                doctorView.in_apell_2.setText(resultado[3])
                doctorView.in_age_2.setText(resultado[5])
                doctorView.in_dir_2.setText(resultado[6])
                doctorView.in_number_2.setText(resultado[7])
                doctorView.in_mail_2.setText(resultado[8])
                sexo = resultado[4]
                if sexo == "Masculino":
                    doctorView.btn_m_2.setChecked(True)
                if sexo == "Femenino":
                    doctorView.btn_f_2.setChecked(True)
                pixmap1 = QPixmap()
                pixmap1.loadFromData(resultado[9])
                doctorView.foto_2.setPixmap(pixmap1)
            widget.addWidget(doctorView)
            widget.setCurrentIndex(widget.currentIndex()+1)
            
            self.hide()
        
 
class ImagePopup(QDialog):
    def __init__(self):
        super().__init__()

        self.image_label = QLabel(self)
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)

        layout = QtWidgets.QVBoxLayout()
        layout.addWidget(self.image_label)
        self.setLayout(layout)

        self.setWindowTitle('Vista completa de la imagen')
        self.center()
        self.setFixedSize(650, 650)
        
    def show_image(self, pixmap):
        self.image_label.setPixmap(pixmap)
        self.image_label.setScaledContents(True)
        self.exec_()  # Muestra la ventana emergente como modal

    def center(self):
        qr = self.frameGeometry()
        cp = QtWidgets.QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

class Ui_placas(QMainWindow):
    def __init__(self,id_user):
        super(Ui_placas, self).__init__()
        loadUi("interfaces/placas.ui", self)
        self.id_user = id_user
        self.btn_agg.clicked.connect(self.addplacas)
        self.btn_edit.clicked.connect(self.editar)
        self.btn_buscar.clicked.connect(self.searchData)
        self.btn_clear.clicked.connect(self.clearInputs)
        self.btn_clear_2.clicked.connect(self.clearInputs_2)
        self.actionSalir.triggered.connect(self.salir)
        self.btn_import.clicked.connect(self.addPhoto)
        self.actionVolver_al_menu_principal.triggered.connect(self.back_menu)
        self.showMaximized()
        
        self.btn_buscar_2.clicked.connect(self.buscarDatos)
        self.img1.mousePressEvent = lambda event: self.show_image_popup(self.img1.pixmap())
        self.img2.mousePressEvent = lambda event: self.show_image_popup(self.img2.pixmap())
        self.img3.mousePressEvent = lambda event: self.show_image_popup(self.img3.pixmap())
        self.img4.mousePressEvent = lambda event: self.show_image_popup(self.img4.pixmap())
        self.img5.mousePressEvent = lambda event: self.show_image_popup(self.img5.pixmap())
        self.img6.mousePressEvent = lambda event: self.show_image_popup(self.img6.pixmap())
        self.img7.mousePressEvent = lambda event: self.show_image_popup(self.img7.pixmap())
        self.img8.mousePressEvent = lambda event: self.show_image_popup(self.img8.pixmap())
        self.img9.mousePressEvent = lambda event: self.show_image_popup(self.img9.pixmap())
        self.img10.mousePressEvent = lambda event: self.show_image_popup(self.img10.pixmap())
        self.usuario =None
        self.verifytipoUser()   
        
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?",(self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser =="Administrador":
                self.usuario = "Administrador"
            elif tipoUser=="Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontro ningun tipo")
                
    def editar(self):
        cedula = self.in_busqueda.text()
        foto_pixmap1  =self.img1.pixmap()
        foto_pixmap2  =self.img2.pixmap()
        foto_pixmap3  =self.img3.pixmap()
        foto_pixmap4  =self.img4.pixmap()
        foto_pixmap5  =self.img5.pixmap()
        if foto_pixmap1 is None or foto_pixmap2 is None or foto_pixmap3 is None or foto_pixmap4 is None or foto_pixmap5 is None:
            QMessageBox.warning(self,"Advertencia","Debes importar 5 imagenes antes de guardar")
            return
        if len(cedula) <=0 :
             QMessageBox.warning(self,"Advertencia","Debes ingresar la cedula para almacenar las placas")
             return
        else:
            try:
                foto1_image = foto_pixmap1.toImage()
                foto2_image = foto_pixmap2.toImage()
                foto3_image = foto_pixmap3.toImage()
                foto4_image = foto_pixmap4.toImage()
                foto5_image = foto_pixmap5.toImage()

                # Convierte cada imagen a un formato de bytes (por ejemplo, PNG)
                foto1_bytes = QByteArray()
                buffer1 = QBuffer(foto1_bytes)
                buffer1.open(QIODevice.WriteOnly)
                foto1_image.save(buffer1, "PNG")
                foto1_byte = buffer1.data()
                buffer1.close()

                foto2_bytes = QByteArray()
                buffer2 = QBuffer(foto2_bytes)
                buffer2.open(QIODevice.WriteOnly)
                foto2_image.save(buffer2, "PNG")
                foto2_byte = buffer2.data()
                buffer2.close()

                foto3_bytes = QByteArray()
                buffer3 = QBuffer(foto3_bytes)
                buffer3.open(QIODevice.WriteOnly)
                foto3_image.save(buffer3, "PNG")
                foto3_byte = buffer3.data()
                buffer3.close()
                
                foto4_bytes = QByteArray()
                buffer4 = QBuffer(foto4_bytes)
                buffer4.open(QIODevice.WriteOnly)
                foto4_image.save(buffer4, "PNG")
                foto4_byte = buffer4.data()
                buffer4.close()
                
                foto5_bytes = QByteArray()
                buffer5 = QBuffer(foto5_bytes)
                buffer5.open(QIODevice.WriteOnly)
                foto5_image.save(buffer5, "PNG")
                foto5_byte = buffer5.data()
                buffer5.close()
                
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                reply = QMessageBox.question(
                    self,
                    'Confirmación',
                    '¿Desea editar las placas ya guardadas?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                if reply == QMessageBox.Yes:
                    cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula = ?", (cedula,))
                    
                    cursor.execute("UPDATE Pacientes SET Placa1 = ?, Placa2 = ? ,Placa3 = ?, Placa4 = ?, Placa5 = ? WHERE Cedula = ?",
                    (foto1_byte, foto2_byte, foto3_byte, foto4_byte, foto5_byte ,cedula ))
                    QMessageBox.information(self, "Exito", "Datos Guardados Correctamente ")
                    conexion.commit()
                    conexion.close()
                    self.clearInputs()
                return
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Error", "Error al actualizar los datos en la base de datos: " + str(e))
    def back_menu(self):
        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)

            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")

            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()

    def show_image_popup(self, pixmap):
        if pixmap:
            
            # Crea una nueva ventana emergente y muestra la imagen
            image_popup = ImagePopup()
            image_popup.show_image(pixmap)

            # Conecta la señal accepted de la ventana emergente para habilitar la ventana principal nuevamente
            image_popup.accepted.connect(self.enable_main_window)

    def enable_main_window(self):
        # Habilita la ventana principal cuando se cierra la ventana emergente
        self.setEnabled(True)
            
    def salir(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()
       
    def clearInputs(self):
        self.in_busqueda.clear()
        self.in_name.clear()
        self.in_apell.clear()
        self.img1.clear()
        self.img2.clear()
        self.img3.clear()
        self.img4.clear()
        self.img5.clear()
        
    def clearInputs_2(self):
        self.in_busqueda_2.clear()
        self.in_apell_2.clear()
        self.in_name_2.clear()
        self.img6.clear()
        self.img7.clear()
        self.img8.clear()
        self.img9.clear()
        self.img10.clear()
        
    def addplacas(self):
        cedula = self.in_busqueda.text()
        foto_pixmap1 = self.img1.pixmap()
        foto_pixmap2 = self.img2.pixmap()
        foto_pixmap3 = self.img3.pixmap()
        foto_pixmap4 = self.img4.pixmap()
        foto_pixmap5 = self.img5.pixmap()

        if foto_pixmap1 is None or foto_pixmap2 is None or foto_pixmap3 is None or foto_pixmap4 is None or foto_pixmap5 is None:
            QMessageBox.warning(self, "Advertencia", "Debes importar 5 imágenes antes de guardar")
            return

        if len(cedula) <= 0:
            QMessageBox.warning(self, "Advertencia", "Debes ingresar la cédula para almacenar las placas")
            return

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            cursor.execute("SELECT Placa1, Placa2, Placa3, Placa4, Placa5 FROM Pacientes WHERE Cedula = ?", (cedula,))
            placas_actuales = cursor.fetchone()

            if any(placa is not None for placa in placas_actuales):
                QMessageBox.critical(self, "Error", "El paciente ya tiene placas registradas.")
            else:
                # Convierte cada imagen a un formato de bytes (por ejemplo, PNG)
                foto1_bytes = QByteArray()
                buffer1 = QBuffer(foto1_bytes)
                buffer1.open(QIODevice.WriteOnly)
                foto_pixmap1.toImage().save(buffer1, "PNG")
                foto1_byte = buffer1.data()
                buffer1.close()

                foto2_bytes = QByteArray()
                buffer2 = QBuffer(foto2_bytes)
                buffer2.open(QIODevice.WriteOnly)
                foto_pixmap2.toImage().save(buffer2, "PNG")
                foto2_byte = buffer2.data()
                buffer2.close()

                foto3_bytes = QByteArray()
                buffer3 = QBuffer(foto3_bytes)
                buffer3.open(QIODevice.WriteOnly)
                foto_pixmap3.toImage().save(buffer3, "PNG")
                foto3_byte = buffer3.data()
                buffer3.close()

                foto4_bytes = QByteArray()
                buffer4 = QBuffer(foto4_bytes)
                buffer4.open(QIODevice.WriteOnly)
                foto_pixmap4.toImage().save(buffer4, "PNG")
                foto4_byte = buffer4.data()
                buffer4.close()

                foto5_bytes = QByteArray()
                buffer5 = QBuffer(foto5_bytes)
                buffer5.open(QIODevice.WriteOnly)
                foto_pixmap5.toImage().save(buffer5, "PNG")
                foto5_byte = buffer5.data()
                buffer5.close()

                cursor.execute("UPDATE Pacientes SET Placa1 = ?, Placa2 = ?, Placa3 = ?, Placa4 = ?, Placa5 = ? WHERE Cedula = ?",
                            (foto1_byte, foto2_byte, foto3_byte, foto4_byte, foto5_byte, cedula))
                QMessageBox.information(self, "Éxito", "Datos Guardados Correctamente ")
                conexion.commit()

            conexion.close()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", "Error al actualizar los datos en la base de datos: " + str(e))

    def searchData(self):
        idUser = self.id_user
        cedula = self.in_busqueda.text()

        if len(cedula) <= 0:
            QMessageBox.warning(self, "Error", "Ingrese una cedula")
        else:
            try:
                conexion = sqlite3.connect('./interfaces/database.db')
                cursor = conexion.cursor()

                # Verificar el rol del usuario
                if self.usuario == "Administrador" or self.usuario == "Usuario":
                    # Si es un administrador o usuario, puede buscar cualquier usuario
                    cursor.execute("SELECT Nombre, Apellido FROM Pacientes WHERE Cedula = ?", (cedula,))
                elif self.usuario == "Doctor":
                    # Si es un doctor, solo puede buscar pacientes asociados a su ID de usuario
                    cursor.execute("SELECT Nombre, Apellido FROM Pacientes WHERE Cedula = ? AND ID_user = ?", (cedula, idUser))
                else:
                    # Manejar el caso en el que el tipo de usuario no sea reconocido
                    QMessageBox.critical(self, "Error", "Rol de usuario no reconocido")
                    return

                resultado = cursor.fetchone()
                conexion.close()
                if resultado :
                    nombre_paciente ,apellido_paciente = resultado
                    self.in_name.setText(nombre_paciente)
                    self.in_apell.setText(apellido_paciente)
                    pass
                else:
                    QMessageBox.critical(self,"Error","El paciente no fue encontrado")
            except sqlite3.Error as error:
                QMessageBox.critical(self, "Error", f"Error al buscar paciente: {str(error)}")

    def addPhoto(self):
            filenames, _ = QFileDialog.getOpenFileNames(self, "Seleccionar imágenes", "", "Archivos de imagen (*.png *.jpg *.bmp *.jpeg *.JFIF)")
                
            if len(filenames) >= 5:
                    pixmap1 = QPixmap(filenames[0])
                    pixmap2 = QPixmap(filenames[1])
                    pixmap3 = QPixmap(filenames[2])
                    pixmap4 = QPixmap(filenames[3])
                    pixmap5 = QPixmap(filenames[4])
                    
                    self.img1.setPixmap(pixmap1)    
                    self.img2.setPixmap(pixmap2)
                    self.img3.setPixmap(pixmap3)
                    self.img4.setPixmap(pixmap4)
                    self.img5.setPixmap(pixmap5)
            else:
                    QMessageBox.information(self,"Imagenes","Por favor,Selecciona una imagen")    
    
    def buscarDatos(self):
        cedula = self.in_busqueda_2.text()
        idUser = self.id_user
        if len(cedula)<= 0:
         QMessageBox.warning(self,"Advertencia","Ingrese una Cédula")
        else:
            try:
                conexion = sqlite3.connect('./interfaces/database.db')
                cursor =conexion.cursor()
                # Verificar el rol del usuario
                if self.usuario == "Administrador" or self.usuario == "Usuario":
                    # Si es un administrador o usuario, puede buscar cualquier usuario
                    cursor.execute("SELECT Nombre , Apellido , Placa1, Placa2, Placa3, Placa4, Placa5 FROM Pacientes WHERE Cedula = ?", (cedula,))
                elif self.usuario == "Doctor":
                    # Si es un doctor, solo puede buscar pacientes asociados a su ID de usuario
                    cursor.execute("SELECT Nombre , Apellido , Placa1, Placa2, Placa3, Placa4, Placa5 FROM Pacientes WHERE Cedula = ? AND ID_user = ?", (cedula, idUser))
                else:
                    # Manejar el caso en el que el tipo de usuario no sea reconocido
                    QMessageBox.critical(self, "Error", "Rol de usuario no reconocido")
                    return

                resultado = cursor.fetchone()
                if resultado:
                    nombre_paciente , apellido_paciente ,placa1 , placa2,placa3, placa4, placa5 = resultado
                    self.in_name_2.setText(nombre_paciente)
                    self.in_apell_2.setText(apellido_paciente)
                    
                    pixmap1 = QPixmap()
                    pixmap1.loadFromData(placa1)
                    self.img6.setPixmap(pixmap1)
                    
                    pixmap2 = QPixmap()
                    pixmap2.loadFromData(placa2)
                    self.img7.setPixmap(pixmap2)
                    
                    pixmap3 = QPixmap()
                    pixmap3.loadFromData(placa3)
                    self.img8.setPixmap(pixmap3)
                    
                    pixmap4 = QPixmap()
                    pixmap4.loadFromData(placa4)
                    self.img9.setPixmap(pixmap4)
                    
                    pixmap5 = QPixmap()
                    pixmap5.loadFromData(placa5)
                    self.img10.setPixmap(pixmap5)
                    
                else:
                    QMessageBox.information(self,"Eror","No se encuentra datos")
                    return
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Error", "Error al actualizar los datos en la base de datos: " + str(e))                   
    def searchAll(self):
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            idUser = self.id_user
            
            # Ejecuta una consulta para obtener los datos de los pacientes y ordenar por Fecha_Cita descendente
            cursor.execute("SELECT Cedula, Nombre, Apellido, Placa1 , Placa2 , Placa3, Placa4, Placa5   FROM Pacientes WHERE ID_user = ? ", (idUser,))
            tabla_placas = cursor.fetchall()
            conexion.close()

            # Filtrar las filas que no contengan ningún valor 'None'
            filas_filtradas = [paciente for paciente in tabla_placas if None not in paciente]

            # Limpiar la tabla existente si es necesario
            self.tabla_placas.clearContents()

            # Establecer el número de filas y columnas en la tabla
            self.tabla_placas.setRowCount(len(filas_filtradas))
            self.tabla_placas.setColumnCount(len(filas_filtradas[0]))

            # Agregar los datos a la tabla
            for row, paciente in enumerate(filas_filtradas):
                for column, value in enumerate(paciente):
                    if column == 6:  
                        if value == 1:
                           
                            item = QTableWidgetItem("Imágenes disponibles")
                        else:
                               item = QTableWidgetItem("No hay imágenes")
                    elif column in [3, 4, 5]:  # Supongamos que las columnas 3, 4 y 5 contienen las rutas de las imágenes
                        if value is not None and os.path.exists(value):                           
                            pixmap = QPixmap(value)
                            if not pixmap.isNull():
                                item = QTableWidgetItem()
                                item.setData(Qt.DecorationRole, pixmap)
                            else:
                               
                                item = QTableWidgetItem("Imagen dañada")
                        else:
                        
                            item = QTableWidgetItem("Imagen faltante")
                    else:
                        item = QTableWidgetItem(str(value))
                    self.tabla_placas.setItem(row, column, item)
        except:
            QMessageBox.critical(self, "Error", "No hay pacientes con placas actualmente.")
            
    def searchForDelete(self):
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            idUser = self.id_user
            cedula = self.in_busqueda_3.text()
            if len(cedula)<= 0:
                QMessageBox.critical(self,"Error","Ingrese la cedula primeramente")
                return
            else:
                
                # Ejecuta una consulta para obtener los datos de los pacientes y ordenar por Fecha_Cita descendente
                cursor.execute("SELECT Cedula, Nombre, Apellido, Placa1 , Placa2 , Placa3, Placa4, Placa5   FROM Pacientes WHERE ID_user = ? AND Cedula = ? ", (idUser,cedula ))
                tabla_placas = cursor.fetchall()
                conexion.close()

                # Filtrar las filas que no contengan ningún valor 'None'
                filas_filtradas = [paciente for paciente in tabla_placas if None not in paciente]

                # Limpiar la tabla existente si es necesario
                self.tabla_delete.clearContents()

                # Establecer el número de filas y columnas en la tabla
                self.tabla_delete.setRowCount(len(filas_filtradas))
                self.tabla_delete.setColumnCount(len(filas_filtradas[0]))

                # Agregar los datos a la tabla
                for row, paciente in enumerate(filas_filtradas):
                    for column, value in enumerate(paciente):
                        if column == 6:  # Supongamos que la columna 6 es "ImagenesDisponibles"
                            if value == 1:
                                # El paciente tiene imágenes, muestra un mensaje o icono personalizado
                                item = QTableWidgetItem("Imágenes disponibles")
                            else:
                                # El paciente no tiene imágenes, muestra un mensaje o icono personalizado
                                item = QTableWidgetItem("No hay imágenes")
                        elif column in [3, 4, 5]:  # Supongamos que las columnas 3, 4 y 5 contienen las rutas de las imágenes
                            if value is not None and os.path.exists(value):
                                pixmap = QPixmap(value)
                                if not pixmap.isNull():
                                    item = QTableWidgetItem()
                                    item.setData(Qt.DecorationRole, pixmap)
                                else:
                                    # La imagen está dañada
                                    # Mostrar un ícono de error en lugar de la imagen
                                    item = QTableWidgetItem("Imagen dañada")
                            else:
                                # La imagen está ausente
                                # Mostrar un ícono de imagen faltante en lugar de la imagen
                                item = QTableWidgetItem("Imagen faltante")
                        else:
                            item = QTableWidgetItem(str(value))
                        self.tabla_delete.setItem(row, column, item)
        except:
            QMessageBox.critical(self, "Error", "El paciente no tiene placas actualmente.")
            
    def DeletePlaca(self):
        try:
            cedula = self.in_busqueda.text()
    
            if len(cedula) == 0:
                QMessageBox.critical(self, "Error", "Ingrese una cédula")
            else:
                Placa1 = None
                Placa2 = None
                Placa3 = None
                Placa4 = None
                Placa5 = None
                conexion = sqlite3.connect('interfaces/database.db')
                cursor = conexion.cursor()
                cursor.execute("UPDATE Pacientes SET Placa1 = ?, Placa2 = ? , Placa3= ?, Placa4 = ?, Placa5 = ? WHERE Cedula = ?",
                        (Placa1, Placa2, Placa3, Placa4, Placa5, cedula))
                conexion.commit()
                conexion.close()
        
        # Eliminación exitosa, muestra un mensaje y realiza otras acciones si es necesario
                QMessageBox.information(self, "Realizado", "Las placas ha sido eliminada correctamente")
                self.tabla_delete.clearContents()
                self.in_busqueda_3.clear()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", "Error al eliminar las placas  de la base de datos: " + str(e))
        pass
    
    
class CalculoDivisaThread(QThread):
    finished = pyqtSignal(float)

    def __init__(self, parent=None):
        super(CalculoDivisaThread, self).__init__(parent)
        self.dolar = 1  # Puedes ajustar el valor predeterminado

    def run(self):
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

        url = 'https://www.bcv.org.ve'
        try:
            response = requests.get(url, verify=False)
            soup = BeautifulSoup(response.text, 'html.parser')
            div_dolar = soup.find('div', id='dolar')
            divisa = div_dolar.find('strong').text
            divisa_limpia = divisa.replace(' ', '').replace(',', '.')
            valor_numerico = float(divisa_limpia)
        except requests.ConnectionError:
            # Si no hay conexión a Internet, extraer el valor de la base de datos
            valor_numerico = self.obtener_valor_bd()

        operacion = float(self.dolar)
        bolivares = operacion * valor_numerico
        suma_formateada = "{:.2f}".format(bolivares).replace(".", ",")

        self.finished.emit(float(suma_formateada.replace(',', '.')))

    def obtener_valor_bd(self):
        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()
        cursor.execute("SELECT valorDola FROM Valor")
        valor_dolar = cursor.fetchone()[0]
        conexion.close()
        return valor_dolar
        
class historiaMenu(QMainWindow):
    def __init__(self ,id_user):
        super(historiaMenu, self).__init__()
        loadUi("interfaces/History.ui", self)
        self.id_user = id_user
        self.valor_numerico = None
        self.btn_buscar.clicked.connect(self.Searchdata)
        self.btn_agg.clicked.connect(self.AddPacient)
        self.btn_clear.clicked.connect(self.clearInputs)
        self.btn_edit.clicked.connect(self.UpdateData)
        self.actionSalir.triggered.connect(self.salir)
        self.actionVolver_al_menu_principal.triggered.connect(self.back_menu)
        self.btn_agg_2.clicked.connect(self.addInformation)
        self.btn_clear_2.clicked.connect(self.clearData)
        self.btn_agg_3.clicked.connect(self.addDiagnostico)
        self.btn_edit_2.clicked.connect(self.addInformation)
        self.btn_edit_3.clicked.connect(self.addDiagnostico)
        self.btn_clear_3.clicked.connect(self.clearDiag)
        self.btn_agg_4.clicked.connect(self.aggTrata)
        self.btn_edit_4.clicked.connect(self.editTrata)
        self.btn_clear_4.clicked.connect(self.clearTrata)
        self.fecha_hora_actualizadas = False
        self.fecha_actualizada = False
        self.tabWidget.currentChanged.connect(self.actualizar_fecha_hora_diagnostico)
        self.tabWidget.currentChanged.connect(self.actualizar_fecha)
        self.showMaximized()
       
        self.usuario =None
        self.verifytipoUser()
        self.verifyUsuario()
        self.tratamientos = {
            "Triaje": ["Consulta e Historia Clínica sin informe", "Consulta e Historia Clínica con informe"],
            "Periodoncia": ["Tartectomía y pulido simple (1 sesión)", "Tartectomía y pulido simple (2-3 sesiones)","Aplicación tópica de fluór","Cirguia periodontal (por cuadrante)"],
            "Blanqueamiento": ["Blanqueamiento intrapulpar", "Blanquemaineto maxilar superior e inferior (2 sesiones de 20 min c/u)"],
            "Operatoria": ["Obturaciones provisionales","Obturaciones con Amalgama","Obturaciones con vidrio ionomerico pequeña","Obturaciones con vidrio ionomerico grande","Obturaciones con resina fotocurada"],
            "Endodoncia": ["Pulpotomías formocreasoladas","Emergencias Endodontica","Tratamiento endodontico monoradicular","Tratamiento endodontico biradicular","Tratamiento endodontico multiradicular","Desobturación conductos"],
            "Radiografias Periaciales": ["Adultos e infantes"],
            "Cirugias": ["Exodoncia simple","Exodoncia quirurgica","Exodoncia de dientes temporales","Exodoncia de corales erupcionadas/incluidas"],
            "Protesis": ["Coronas provisionales por unidad","Muñon artificial monoradicular","Muñon artificial multiradicular","Incrustacion resina/metálica","Unidad de corona meta-porcelana","Cementado de protesis fija"],
            "Protesis removibles metalicas y/o acrilicas": ["1 a 3 unidades","4 a 6 unidades","7 a 12 unidades","Unidadad adicional","Ganchos contorneados retentativas acrilicas c/u","Reparaciones protesis acrilicas y/oo agregar un diente a la protesis"],
            "Protesis totales": ["Dentadura superior o inferior (incluye controles post-inatalción) c/u"],
            "Implantes dentales": ["Honorarios cirujano por implante","Implante y aditamientos","Injertos óseos (1cc)","PRF (incluye bionalista y extraccion de sangre + centrifugado)","Corona metal porcelana sobre implante","DPR acrilica"],
        }

        self.combo_honorario = [self.findChild(QtWidgets.QComboBox, f"c_{i}") for i in range(6)]
        self.combo_tratamiento = [self.findChild(QtWidgets.QComboBox, f"t_{i}") for i in range(6)]
        self.monto_dola = [self.findChild(QtWidgets.QLineEdit, f"monto_dola_{i+1}") for i in range(6)]
        self.monto_bs = [self.findChild(QtWidgets.QLineEdit, f"monto_bs_{i+1}") for i in range(6)]
        for combo in self.combo_honorario:
            combo.addItem("Seleccione el tipo de honorario")
            combo.addItems(list(self.tratamientos.keys()))
            combo.currentTextChanged.connect(self.loadTratamientos)

        for i, combo in enumerate(self.combo_tratamiento):
            combo.addItem("Seleccione el tipo de tratamiento")
            combo.currentTextChanged.connect(lambda _, index=i: self.update_monto(index))
            
            
        # Iniciar el hilo para calcular la divisa
        self.calculo_divisa_thread = CalculoDivisaThread()
        self.calculo_divisa_thread.finished.connect(self.on_calculo_divisa_finished)
        self.calculo_divisa_thread.start()
        
    
    def on_calculo_divisa_finished(self, resultado):
        # Este método se llama cuando el hilo ha terminado de calcular la divisa
        self.valor_numerico = resultado
        print(f"El valor_numerico al iniciar el programa es: {self.valor_numerico}")

    def calcularDivisas(self,dolar):
        bolivar = self.valor_numerico
        operacion = bolivar * dolar
        return operacion, bolivar
    def verifytipoUser(self):
        conexion = sqlite3.connect("./interfaces/database.db")
        cursor = conexion.cursor()
        cursor.execute("SELECT Tipo FROM Users WHERE ID=?",(self.id_user,))
        resultado = cursor.fetchone()
        if resultado:
            tipoUser = resultado[0]
            if tipoUser == "Doctor":
                self.usuario = "Doctor"
            elif tipoUser =="Administrador":
                self.usuario = "Administrador"
            elif tipoUser=="Usuario":
                self.usuario = "Usuario"
            else:
                print("No se encontro ningun tipo")
    def verifyUsuario(self):
        if self.usuario== "Usuario":
            
            self.btn_agg_2.setEnabled(False)
            self.btn_edit_2.setEnabled(False)
            self.btn_clear_2.setEnabled(False)
            
            self.btn_agg_3.setEnabled(False)
            self.btn_edit_3.setEnabled(False)
            self.btn_clear_3.setEnabled(False)
            
            self.btn_agg_4.setEnabled(False)
            self.btn_edit_4.setEnabled(False)
            self.btn_clear_4.setEnabled(False)
            
    def actualizar_fecha_hora_diagnostico(self):
        # Solo actualiza los campos de fecha y hora si no se han actualizado previamente
        if not self.fecha_hora_actualizadas:
            fecha_actual = QDate.currentDate()
            hora_actual = QTime.currentTime()

            self.fecha.setDate(fecha_actual)
            self.hora.setTime(hora_actual)

            # Marca que los campos se han actualizado
            self.fecha_hora_actualizadas = True
            
    def actualizar_fecha(self):
        # Solo actualiza los campos de fecha y hora si no se han actualizado previamente
        if not self.fecha_actualizada:
            fecha_actual = QDate.currentDate()

            self.fecha_2.setDate(fecha_actual)

            # Marca que los campos se han actualizado
            self.fecha_actualizada = True
            
    def clearTrata(self):
        self.in_name_4.clear()
        self.in_apell_4.clear()
        self.monto_dola_1.clear()
        self.monto_bs_1.clear()
        self.monto_dola_2.clear()
        self.monto_bs_2.clear()
        self.monto_dola_3.clear()
        self.monto_bs_3.clear()
        self.monto_dola_4.clear()
        self.monto_bs_4.clear()
        self.monto_dola_5.clear()
        self.monto_bs_5.clear()
        self.monto_dola_6.clear()
        self.monto_bs_6.clear()
        self.totalBs.clear()
        self.totaldola.clear()
        

    def loadTratamientos(self):
        sender = self.sender()
        selected_honorario = sender.currentText()

        if selected_honorario == "Seleccione el tipo de honorario":
            return

        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()

        index = self.combo_honorario.index(sender)

        tratamientos_honorario = self.tratamientos.get(selected_honorario, [])
        linked_combo = self.combo_tratamiento[index]
        linked_combo.clear()

        for tratamiento in tratamientos_honorario:
            cursor.execute("SELECT monto FROM Trata WHERE tipo_tratamiento = ? AND tratamiento = ?", (selected_honorario, tratamiento))
            monto = cursor.fetchone()
            if monto:
                linked_combo.addItem(tratamiento)

    def update_monto(self, index):
        selected_honorario = self.combo_honorario[index].currentText()
        selected_tratamiento = self.combo_tratamiento[index].currentText()

        conexion = sqlite3.connect('interfaces/database.db')
        cursor = conexion.cursor()

        cursor.execute("SELECT monto FROM Trata WHERE tipo_tratamiento = ? AND tratamiento = ?", (selected_honorario, selected_tratamiento))
        monto = cursor.fetchone()

        if monto:
            self.monto_dola[index].setText(str(monto[0]))
           
            dolares = monto[0]
            resultado_divisa = self.calcularDivisas(dolares)
            if resultado_divisa:
                bolivares = resultado_divisa[0]  # Obtener el primer elemento de la tupla
                bolivares_formateados = "{:.1f}".format(bolivares)  #
                self.monto_bs[index].setText(str(bolivares_formateados))
                self.totalPrecio()
            else:
                # Manejar el caso en el que no se pueda obtener la información de la divisa
                self.monto_bs[index].setText("Error al obtener la tasa de cambio")
        else:
            # Manejar el caso en el que no se encuentre el monto en la base de datos
            self.monto_bs[index].setText("Monto no encontrado en la base de datos")

           
    def totalPrecio(self):
        # Calcula el total sumando los valores flotantes de los QLineEdits
        totalBs = 0.0
        totalUsd= 0.0

        # Asegúrate de manejar las conversiones a flotante correctamente y verifica que las cadenas no estén vacías
        if self.monto_bs_1.text():
            totalBs += float(self.monto_bs_1.text().replace(',', '.'))

        if self.monto_bs_2.text():
            totalBs += float(self.monto_bs_2.text().replace(',', '.'))

        if self.monto_bs_3.text():
            totalBs += float(self.monto_bs_3.text().replace(',', '.'))
        if self.monto_bs_4.text():
            totalBs += float(self.monto_bs_4.text().replace(',', '.'))
        if self.monto_bs_5.text():
            totalBs += float(self.monto_bs_5.text().replace(',', '.'))
        if self.monto_bs_6.text():
            totalBs += float(self.monto_bs_6.text().replace(',', '.'))

        if self.monto_dola_1.text():
            totalUsd += float(self.monto_dola_1.text().replace(',', '.'))

        if self.monto_dola_2.text():
            totalUsd += float(self.monto_dola_2.text().replace(',', '.'))

        if self.monto_dola_3.text():
            totalUsd += float(self.monto_dola_3.text().replace(',', '.'))
        if self.monto_dola_4.text():
            totalUsd += float(self.monto_dola_4.text().replace(',', '.'))
        if self.monto_dola_5.text():
            totalUsd += float(self.monto_dola_5.text().replace(',', '.'))
        if self.monto_dola_6.text():
            totalUsd += float(self.monto_dola_6.text().replace(',', '.'))

        # Formatea el número con dos decimales
        totalBs_formateado = f'{totalBs:.2f}'
        totalUsd_formateado = f'{totalUsd:.2f}'
    
        # Asigna el valor formateado al QLineEdit
        self.totalBs.setText(totalBs_formateado)
        self.totaldola.setText(totalUsd_formateado)
    
        
    def clearDiag(self):
        self.in_name_3.clear()
        self.in_apell_3.clear()
        self.hora.clear()
        self.fecha.clear()
        self.diag.clear()
        
    def addDiagnostico(self):
        cedula = self.in_busqueda.text()
        diag = self.diag.toPlainText()
        fecha = self.fecha.date()
        hora = self.hora.time()
        fechaToString = fecha.toString('yyyy-MM-dd')
        horatoString = hora.toString('hh:mm:ss')
        print (fechaToString)
        print (horatoString)
        if not cedula :
            QMessageBox.warning(self,"Error","Por favor ingrese la cedula en el menu de registro ")
            return
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
                        # Verificar si ya existe un paciente con la misma cédula
            cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula = ?", (cedula,))
            cursor.execute("UPDATE Pacientes SET Diagnotico=?, Fecha_Diagnotico=?, Hora_Diagnostico=?  WHERE Cedula=?", (
                diag,fechaToString,horatoString, cedula ))
            conexion.commit()
            QMessageBox.information(self, "Éxito", "Informacion registrada correctamente.")

            # Cierra la conexión con la base de datos
            conexion.close()
    
        # Actualizar los registros en la base de datos
        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al registrar el paciente: {str(error)}")
    def clearData(self):
        self.hiper_control.clear()
        self.diabetes_control.clear()
        self.coagualcion_control.clear()
        self.ln_data.clear()
        self.btn_si.setChecked(False)
        self.btn_si_4.setChecked(False)
        self.btn_no_4.setChecked(False)
        self.btn_si_3.setChecked(False)
        self.btn_no_3.setChecked(False)
        self.btn_no.setChecked(False)
        self.ln_alergias.clear()
        
    def salir(self):
        reply = QMessageBox.question(
            self,
            'Confirmación',
            '¿Desea Salir?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            QApplication.quit()

    def addInformation(self):
        hipertenso = None
        diabetes = None
        coagulacion = None
        cedula = self.in_busqueda.text()
        hipertenso_control = self.hiper_control.toPlainText()
        diabetes_control = self.diabetes_control.toPlainText()
        coagualcion_control = self.coagualcion_control.toPlainText()

        if not cedula:
            QMessageBox.warning(self, "Error", "Por favor ingrese la cedula en el menú de registro.")
            return

        if self.btn_si.isChecked():
            hipertenso = "Si"
        if self.btn_no.isChecked():
            hipertenso = "No"

        if self.btn_si_4.isChecked():
            diabetes = "Si"
        if self.btn_no_4.isChecked():
            diabetes = "No"

        if self.btn_si_3.isChecked():
            coagulacion = "Si"
        if self.btn_no_3.isChecked():
            coagulacion = "No"

        Otros = self.ln_data.toPlainText()
        alergias = self.ln_alergias.toPlainText()

        if not Otros:
            QMessageBox.warning(self, "Error", "Por favor ingrese información en el campo 'Otros'.")
            return

        if not alergias:
            QMessageBox.warning(self, "Error", "Por favor ingrese información en el campo 'Alergias'.")
            return

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            
            # Verificar si ya existe un paciente con la misma cédula
            cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula = ?", (cedula,))
            existe_paciente = cursor.fetchone()[0]
            if existe_paciente:
                cursor.execute("UPDATE Pacientes SET Hipertension=?, Diabates=?, Coagualcion=?, Otros=?, Alergias=?, diabate_Data=?, hipertension_Data=?, Coagualcion_Data=?  WHERE Cedula=?", (
                    hipertenso, diabetes, coagulacion, Otros, alergias, diabetes_control, hipertenso_control, coagualcion_control, cedula))
                conexion.commit()
                QMessageBox.information(self, "Éxito", "Información registrada correctamente.")
            else:
                QMessageBox.warning(self, "Advertencia", "No se encontró un paciente con la cédula proporcionada.")

            # Cierra la conexión con la base de datos
            conexion.close()

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al registrar la información del paciente: {str(error)}")


    def clearInputs(self):
        self.in_cedula.clear()
        self.in_name.clear()
        self.in_apell.clear()
        self.in_age.clear()
        self.in_mail.clear()
        self.in_number.clear()
        self.in_dir.clear()
        self.btn_m.setChecked(False)
        self.btn_f.setChecked(False)
        self.motivo.clear()
        
    def AddPacient(self): 
        idUser = self.id_user
        cedula = self.in_cedula.text()
        nombre = self.in_name.text()
        apellido = self.in_apell.text()
        edad = self.in_age.text()
        mail = self.in_mail.text()
        valor_sexo = None
        if self.btn_m.isChecked():
            valor_sexo = "Masculino"
        if self.btn_f.isChecked():
            valor_sexo = "Femenino"
        telefono = self.in_number.text()
        direccion = self.in_dir.text()
        contexto = self.motivo.toPlainText()
        
        if len(cedula) < 8:
            QMessageBox.critical(self, "Error", "La cedula debe tener mínimo 8 caracteres.")
            return
        
        if valor_sexo is None:
            QMessageBox.critical(self,"Error","Por favor seleccione su sexo")
            return

        if not cedula or not nombre or not apellido or not edad or not valor_sexo or not mail  or not telefono or not direccion or not contexto:
            QMessageBox.critical(self, "Error", "Por favor, complete todos los campos.")
            return

        telefono_pattern = re.compile(r'^\d{11}$')  # Asume que el número de teléfono debe tener 10 dígitos
        mail_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')  # Patrón para validar correo electrónico
        cedula_pattern = re.compile(r'^\d+$')  # Asume que la cédula debe contener solo dígitos
        edad_pattern = re.compile(r'^\d+$')  # Asume que la edad debe contener solo dígitos
        
        if not cedula_pattern.match(cedula):
            QMessageBox.warning(self, "Error", "Ingrese una cédula válida (solo números).")
            return
        if not telefono_pattern.match(telefono):
            QMessageBox.warning(self, "Error", "Ingrese un número de teléfono válido (0000-0000-000).")
            return
        if not edad_pattern.match(edad):
            QMessageBox.warning(self, "Error", "Ingrese una edad válida (solo números).")
            return

        if not mail_pattern.match(mail):
            QMessageBox.warning(self, "Error", "Ingrese una dirección de correo electrónico válida.")
            return
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            # Verificar si ya existe un paciente con la misma cédula
            cursor.execute("SELECT COUNT(*) FROM Pacientes WHERE Cedula = ?", (cedula,))
            existe_paciente = cursor.fetchone()[0]
            if existe_paciente:
                QMessageBox.warning(self, "Advertencia", "Ya existe un paciente con la misma cédula.")
                return
            else:
                cursor.execute("INSERT INTO Pacientes (Cedula, Nombre, Apellido, Edad, Sexo ,Direccion , ID_user ,Telefono, Mail ,Context) VALUES (?, ?, ?, ?, ?, ?, ? , ? ,? , ?)",
                            (cedula, nombre, apellido, edad, valor_sexo , direccion , idUser ,telefono , mail , contexto))

                # Confirmar los cambios en la base de datos
                conexion.commit()

                QMessageBox.information(self, "Éxito", "Paciente registrado correctamente.")

                # Cierra la conexión con la base de datos
                conexion.close()

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al registrar el paciente: {str(error)}")


    def Searchdata(self):
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            idUser = self.id_user
            busqueda = self.in_busqueda.text()

            # Verificar el rol del usuario
            if self.usuario == "Administrador"or self.usuario == "Usuario":
                # Si es un administrador, puede buscar cualquier usuario
                cursor.execute("SELECT Cedula, Nombre, Apellido, Edad, Direccion, Sexo, Telefono, Mail, Context, Hipertension, Diabates, Coagualcion, Otros, Alergias, Diabate_Data, Hipertension_Data, Coagualcion_Data, Diagnotico, Fecha_Diagnotico, Hora_Diagnostico FROM Pacientes WHERE Cedula = ?", (busqueda,))
            elif self.usuario == "Doctor":
                cursor.execute("SELECT Cedula, Nombre, Apellido, Edad, Direccion, Sexo, Telefono, Mail, Context, Hipertension, Diabates, Coagualcion, Otros, Alergias, Diabate_Data, Hipertension_Data, Coagualcion_Data, Diagnotico, Fecha_Diagnotico, Hora_Diagnostico FROM Pacientes WHERE Cedula = ? AND ID_user = ?", (busqueda, idUser))
            else:
                # Manejar el caso en el que el tipo de usuario no sea reconocido
                QMessageBox.critical(self, "Error", "Rol de usuario no reconocido")

            resultado = cursor.fetchone()

            if resultado:
                (Cedula, Nombre, Apellido, Edad, Direccion, Sexo, 
                    Telefono, Mail, Context ,Hipertension, Diabates, Coagualcion, Otros, Alergias, 
                    Diabate_Data , Hipertension_Data , Coagualcion_Data, Diagnotico, Fecha_Diagnotico, Hora_Diagnostico) = resultado
            
                self.in_cedula.setText(Cedula)
                self.in_name.setText(Nombre)
                self.in_apell.setText(Apellido)
                self.in_age.setText(Edad)  
                self.in_mail.setText(Mail)  # Correo
                self.in_dir.setText(Direccion)
                self.in_number.setText(Telefono)  # Prueba##
                self.motivo.setText(Context)
                self.hiper_control.setText(Hipertension_Data)
                self.diabetes_control.setText(Diabate_Data)
                self.coagualcion_control.setText(Coagualcion_Data)
                self.ln_data.setText(Otros)
                self.ln_alergias.setText(Alergias)
                self.in_name_3.setText(Nombre)
                self.in_apell_3.setText(Apellido)
                self.diag.setText(Diagnotico)
                self.in_name_4.setText(Nombre)
                self.in_apell_4.setText(Apellido)
                
                if Hipertension == "Si":
                    self.btn_si.setChecked(True)
                else:
                    self.btn_no.setChecked(True)
                if Diabates =="Si":
                    self.btn_si_4.setChecked(True)
                else:
                    self.btn_no_4.setChecked(True)
                if Coagualcion =="Si":
                    self.btn_si_3.setChecked(True)
                else:
                    self.btn_no_3.setChecked(True)

                # Manejar los botones de radio según el valor de Sexo
                if Sexo == "Masculino":
                    self.btn_m.setChecked(True)
                if Sexo == "Femenino":
                    self.btn_f.setChecked(True)

                # Obtener la fecha y hora de la base de datos
                if Fecha_Diagnotico is not None:
                    fecha_cita = QDate.fromString(Fecha_Diagnotico, 'yyyy-MM-dd')
                    self.fecha.setDate(fecha_cita)
                if Hora_Diagnostico is not None:
                    hora_cita = QTime.fromString(Hora_Diagnostico, 'hh:mm:ss')
                    self.hora.setTime(hora_cita)
                
                self.tabla_pacientes.clearContents()
                self.tabla_pacientes.setRowCount(1)
                self.tabla_pacientes.setColumnCount(8)
                data = [Cedula, Nombre, Apellido, Edad, Sexo, Direccion, Telefono, Mail]
                for column, value in enumerate(data):
                    item = QTableWidgetItem(value)
                    self.tabla_pacientes.setItem(0, column, item)
                    
                # Cargar los tratamientos existentes en las QComboBox
                cursor.execute("SELECT ID_Trata FROM PTrata WHERE Cedula = ?", (busqueda,))
                id_trata = cursor.fetchone()

                if id_trata:
                    cursor.execute(f"SELECT * FROM PTrata WHERE ID_Trata = ?", (id_trata[0],))
                    tratamientos = cursor.fetchone()

                    column_names = [description[0] for description in cursor.description]

                    for i in range(6):
                        tipo_trata_column = f"Tipo_Trata{i + 1}"
                        trata_column = f"Tratamiento{i + 1}"

                        tipo_tratamiento = tratamientos[column_names.index(tipo_trata_column)]
                        tratamiento = tratamientos[column_names.index(trata_column)]

                        self.combo_honorario[i].setCurrentText(tipo_tratamiento)
                        self.combo_tratamiento[i].setCurrentText(tratamiento)
            else:
                # Limpiar la tabla existente si no se encuentra ningún registro
                self.tabla_pacientes.clearContents()
                QMessageBox.warning(self, "Advertencia", "No se ha encontrado ningún registro")

            conexion.close()
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", "Error al consultar la base de datos: " + str(e))

    def UpdateData(self):
        busqueda = self.in_busqueda.text()
        if not busqueda:
            QMessageBox.warning(self, "Error", "Introduzca su cédula")
            return

        try:
            cedula = self.in_cedula.text()
            nombre = self.in_name.text()
            apellido = self.in_apell.text()
            edad = self.in_age.text()
            direccion = self.in_dir.text()
            telefono = self.in_number.text()
            mail = self.in_mail.text()
            valor_sexo = None
            if self.btn_m.isChecked():
                valor_sexo = "Masculino"
            if self.btn_f.isChecked():
                valor_sexo = "Femenino"
            context = self.motivo.toPlainText()

            telefono_pattern = re.compile(r'^\d{11}$')  # Asume que el número de teléfono debe tener 10 dígitos
            mail_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')  # Patrón para validar correo electrónico
            edad_pattern = re.compile(r'^\d+$')  # Asume que la edad debe contener solo dígitos

            if not edad_pattern.match(edad):
                QMessageBox.warning(self, "Error", "Ingrese una edad válida (solo números).")
                return

            if not telefono_pattern.match(telefono):
                QMessageBox.warning(self, "Error", "Ingrese un número de teléfono válido (0000-0000-000).")
                return

            if not mail_pattern.match(mail):
                QMessageBox.warning(self, "Error", "Ingrese una dirección de correo electrónico válida.")
                return

            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            # Actualizar los registros en la base de datos
            cursor.execute("UPDATE Pacientes SET Nombre=?, Apellido=?, Edad=?, Direccion=?, Sexo=? ,Telefono=? ,Mail=? , Context=? WHERE Cedula=?", 
                            (nombre, apellido, edad, direccion, valor_sexo, telefono, mail, context, cedula))
            conexion.commit()

            QMessageBox.information(self, "Información", "Los datos se actualizaron correctamente")

            conexion.close()
            self.Searchdata()

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Error", "Error al actualizar los datos en la base de datos: " + str(e))
    
    def editTrata(self):
        cedula = self.in_busqueda.text()

        # Verificar si la cédula está presente
        if not cedula:
            QMessageBox.warning(self, "Error", "Por favor ingrese la cédula en el menú de registro ")
            return

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            # Verificar si ya existe un registro para el paciente
            cursor.execute("SELECT ID_Trata FROM PTrata WHERE Cedula = ?", (cedula,))
            row = cursor.fetchone()

            if not row:
                QMessageBox.warning(self, "Error", "No hay tratamientos registrados para este paciente.")
                return

            # Si existe, obtener el ID existente
            id_trata = row[0]

            # Obtener los tratamientos existentes del paciente
            cursor.execute(f"SELECT * FROM PTrata WHERE ID_Trata = ?", (id_trata,))
            tratamientos = cursor.fetchone()

            # Verificar si existen tratamientos antes de intentar editar
            if not tratamientos:
                QMessageBox.warning(self, "Error", "No hay tratamientos registrados para este paciente.")
                return

            # Obtener los nombres de las columnas
            column_names = [description[0] for description in cursor.description]

            # Actualizar los tratamientos existentes con los valores seleccionados en las QComboBox
            for i in range(6):
                tipo_trata_column = f"Tipo_Trata{i + 1}"
                trata_column = f"Tratamiento{i + 1}"

                tipo_tratamiento = self.combo_honorario[i].currentText()
                tratamiento = self.combo_tratamiento[i].currentText()

                if tipo_tratamiento and tratamiento:
                    cursor.execute(f"""
                        UPDATE PTrata
                        SET {tipo_trata_column} = ?, {trata_column} = ?
                        WHERE ID_Trata = ?
                    """, (tipo_tratamiento, tratamiento, id_trata))

            # Confirmar los cambios en la base de datos
            conexion.commit()
            QMessageBox.information(self, "Éxito", "Tratamientos editados correctamente.")

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al editar los tratamientos: {str(error)}")

        finally:
            # Cierra la conexión con la base de datos
            conexion.close()

            
    def aggTrata(self):
        # Obtener información del paciente
        cedula = self.in_busqueda.text()
        fecha_tratamiento = self.fecha_2.date().toString('yyyy-MM-dd')

        # Verificar si la cédula está presente
        if not cedula:
            QMessageBox.warning(self, "Error", "Por favor ingrese la cédula en el menú de registro ")
            return

        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()

            # Verificar si ya existe un registro para el paciente
            cursor.execute("SELECT ID_Trata FROM PTrata WHERE Cedula = ?", (cedula,))
            row = cursor.fetchone()

            if row:
                # Si existe, obtener el ID existente
                id_trata = row[0]
            else:
                # Si no existe, insertar un nuevo registro y obtener el ID asignado
                cursor.execute("INSERT INTO PTrata (Cedula) VALUES (?)", (cedula,))
                id_trata = cursor.lastrowid

                # Confirmar los cambios en la base de datos
                conexion.commit()

            # Verificar si ya hay tratamientos registrados para este paciente
            cursor.execute(f"SELECT COUNT(*) FROM PTrata WHERE ID_Trata = ?", (id_trata,))
            tratamientos_registrados = cursor.fetchone()[0]

            if tratamientos_registrados >= 6:
                QMessageBox.warning(self, "Advertencia", "Este paciente ya tiene 6 tratamientos registrados.")
                return

            # Recorrer las combobox y obtener los tratamientos
            for i in range(6):
                honorario = self.combo_honorario[i].currentText()
                tratamiento = self.combo_tratamiento[i].currentText()

                # Verificar si se seleccionó un tratamiento
                if honorario != "Seleccione el tipo de honorario" and tratamiento:
                    # Crear los nombres de las columnas según el tipo de tratamiento
                    tipo_trata_column = f"Tipo_Trata{i + 1}"
                    trata_column = f"Tratamiento{i + 1}"
                    fecha_trata_column = "Fecha_Trata"

                    # Verificar si ya existe un tratamiento para esa columna
                    cursor.execute(f"SELECT {tipo_trata_column}, {trata_column} FROM PTrata WHERE ID_Trata = ?", (id_trata,))
                    existing_data = cursor.fetchone()

                    # Si ya existe, mostrar un mensaje de advertencia y salir del bucle
                    if existing_data[0] or existing_data[1]:
                        QMessageBox.warning(self, "Advertencia", f"Este paciente ya tiene tratamientos registrados. Por favor, edite los ya asignados.")
                        return

                    # Si no existe, actualizar la columna correspondiente
                    cursor.execute(f"""
                        UPDATE PTrata
                        SET {tipo_trata_column} = ?, {trata_column} = ?, {fecha_trata_column} = ?
                        WHERE ID_Trata = ?
                    """, (honorario, tratamiento, fecha_tratamiento, id_trata))

            # Confirmar los cambios en la base de datos
            conexion.commit()
            QMessageBox.information(self, "Éxito", "Tratamientos registrados correctamente.")

            # Cierra la conexión con la base de datos
            conexion.close()

        except sqlite3.Error as error:
            QMessageBox.critical(self, "Error", f"Error al registrar tratamientos: {str(error)}")

    def searchDataForDelete(self):
        try:
            conexion = sqlite3.connect('interfaces/database.db')
            cursor = conexion.cursor()
            idUser = self.id_user
            cedula = self.in_busqueda_delete.text()
            # Ejecuta una consulta para obtener los datos de los pacientes
            cursor.execute("SELECT Cedula, Nombre, Apellido, Edad, Direccion , Sexo  FROM Pacientes WHERE Cedula = ? AND ID_user = ?", (cedula , idUser) )
            
           
            tabla_pacientes = cursor.fetchall()
            conexion.close()

            # Limpiar la tabla existente si es necesario
            self.tabla_pacientes.clearContents()

            # Establecer el número de filas y columnas en la tabla
            self.tabla_pacientes.setRowCount(len(tabla_pacientes))
            self.tabla_pacientes.setColumnCount(len(tabla_pacientes[0]))

            # Agregar los datos a la tabla
            for row, paciente in enumerate(tabla_pacientes):
                for column, value in enumerate(paciente):
                    item = QTableWidgetItem(str(value))
                    self.tabla_pacientes.setItem(row, column, item)
        except:
             QMessageBox.critical(self, "Error", "No hay ningún paciente con esa cedula.")
   
    def back_menu(self):
        
        conexion = sqlite3.connect('interfaces/database.db')
        cursor= conexion.cursor()
        cursor.execute("SELECT Username FROM Users WHERE ID = ?", (self.id_user,))
        
        resultado = cursor.fetchone()
        if resultado :
            nombre_usuario = resultado[0]
            horaActual = datetime.datetime.now().time()
            
            if datetime.time(5, 0, 0) <= horaActual < datetime.time(12, 0, 0):
                textForMenu = f"Buenos días {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(12, 0, 0) <= horaActual < datetime.time(18, 0, 0):
                textForMenu = f"Buenas tardes {nombre_usuario}\n¿Qué deseas hacer hoy?"
            elif datetime.time(18, 0, 0) <= horaActual or horaActual < datetime.time(5, 0, 0):
                textForMenu = f"Buenas noches {nombre_usuario}\n¿Qué deseas hacer hoy?"
            else:
                textForMenu = f"Hola {nombre_usuario}\n¿Qué deseas hacer hoy?"
            menu_principal = MenuPrincipal(self.id_user)
            menu_principal.lb_nombre.setText(textForMenu)

            # Establecer la ventana en modo de pantalla completa
            menu_principal.showMaximized()

            menu_principal.setWindowTitle("Menu Principal")

            # Asegúrate de añadir la ventana al widget después de establecerla en modo de pantalla completa
            widget.addWidget(menu_principal)
            widget.setCurrentIndex(widget.currentIndex() + 1)

            self.close()
       
    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("Clinica")  # Establecer el nombre de la aplicación

    ingreso_usuario = IngresoUsuario()
    
    widget = QStackedWidget()
    widget.addWidget(ingreso_usuario)
    widget.setGeometry(ingreso_usuario.geometry())
    widget.show()
    icon = QIcon("./interfaces/ELEMENTOS GRAFICOS/odontology-outline.png")
    ingreso_usuario.setWindowIcon(icon)
    sys.exit(app.exec_())